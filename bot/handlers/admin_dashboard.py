import datetime
import pandas as pd
import io
from datetime import datetime as dt

from telegram import ReplyKeyboardMarkup
from telegram.ext import ConversationHandler
import logging
from openpyxl.chart import LineChart, BarChart, Reference

from bot.services.database import get_db
from bot.services.auth import require_admin
from bot.handlers.agent import get_lang, _collect_future_obligations
from bot.services.localization import _

logger = logging.getLogger(__name__)


# این فایل توابع اصلی داشبورد ادمین را مدیریت می‌کند:
# - نمایش خلاصهٔ آماری سیستم برای ادمین
# - تولید و دانلود گزارش کامل اکسل


# =======================
# 📈 پرمختللې احصایوي ډشبورډ
# =======================


@require_admin
async def dashboard_stats(update, context):
    # محاسبه و نمایش داشبورد آماری پیشرفتهٔ سیستم برای ادمین
    lang = get_lang(context)
    if lang == "fa":
        await update.message.reply_text("📈 در حال آماده‌سازی داشبورد آماری...")
    else:
        await update.message.reply_text("📈 د احصایوي ډشبورډ چمتو کول...")

    conn = get_db()
    cur = conn.cursor()
    
    # ټول احصایې
    cur.execute("SELECT COUNT(*) FROM agents WHERE is_active = 1")
    active_agents = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM transactions")
    total_transactions = cur.fetchone()[0]
    
    cur.execute("SELECT SUM(amount) FROM transactions WHERE status != 'cancelled'")
    total_amount = cur.fetchone()[0] or 0
    
    cur.execute("SELECT SUM(commission) FROM transactions WHERE status != 'cancelled'")
    total_commission = cur.fetchone()[0] or 0
    
    # ۷ ورځې تېره احصایې
    seven_days_ago = (dt.now() - datetime.timedelta(days=7)).strftime('%Y-%m-%d')
    cur.execute("""
        SELECT COUNT(*) FROM transactions 
        WHERE DATE(created_at) >= ?
    """, (seven_days_ago,))
    last_7_days_transactions = cur.fetchone()[0]
    
    cur.execute("""
        SELECT SUM(amount) FROM transactions 
        WHERE DATE(created_at) >= ? AND status != 'cancelled'
    """, (seven_days_ago,))
    last_7_days_amount = cur.fetchone()[0] or 0
    
    # ۳۰ ورځې تېره احصایې
    thirty_days_ago = (dt.now() - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
    cur.execute("""
        SELECT COUNT(*) FROM transactions 
        WHERE DATE(created_at) >= ?
    """, (thirty_days_ago,))
    last_30_days_transactions = cur.fetchone()[0]
    
    cur.execute("""
        SELECT SUM(amount) FROM transactions 
        WHERE DATE(created_at) >= ? AND status != 'cancelled'
    """, (thirty_days_ago,))
    last_30_days_amount = cur.fetchone()[0] or 0
    
    # پرکارترین عامل‌ها
    cur.execute("""
        SELECT a.name, COUNT(t.id) as transaction_count,
               SUM(t.commission) as total_commission
        FROM agents a
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        WHERE a.is_active = 1
        GROUP BY a.id, a.name
        ORDER BY transaction_count DESC
        LIMIT 3
    """)
    top_agents = cur.fetchall()
    
    # کم‌کارترین عامل‌ها (فعال)
    cur.execute("""
        SELECT a.name, COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        WHERE a.is_active = 1
        GROUP BY a.id, a.name
        ORDER BY transaction_count ASC
        LIMIT 3
    """)
    least_active_agents = cur.fetchall()
    
    # بیشترین درآمدزا
    cur.execute("""
        SELECT a.name, SUM(t.commission) as total_commission,
               COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        WHERE a.is_active = 1
        GROUP BY a.id, a.name
        HAVING total_commission > 0
        ORDER BY total_commission DESC
        LIMIT 3
    """)
    top_earners = cur.fetchall()
    
    conn.close()

    if lang == "fa":
        dashboard = "📈 *داشبورد آماری پیشرفته*\n"
        dashboard += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
        dashboard += "📊 *آمار کلی سیستم:*\n"
        dashboard += f"   👥 عامل‌های فعال: {active_agents}\n"
        dashboard += f"   📦 کل حواله‌ها: {total_transactions:,}\n"
        dashboard += f"   💰 مجموع مبالغ: {total_amount:,.0f} افغانی\n"
        dashboard += f"   💸 مجموع کارمزدها: {total_commission:,.0f} افغانی\n\n"
        dashboard += "📅 *مقایسه دوره‌های زمانی:*\n"
        dashboard += f"   🗓️ ۷ روز گذشته: {last_7_days_transactions:,} حواله ({last_7_days_amount:,.0f} افغانی)\n"
        dashboard += f"   🗓️ ۳۰ روز گذشته: {last_30_days_transactions:,} حواله ({last_30_days_amount:,.0f} افغانی)\n"
        if last_7_days_transactions > 0:
            avg_per_day = last_7_days_transactions / 7
            dashboard += f"   📊 میانگین روزانه (۷ روز): {avg_per_day:.1f} حواله\n"
        dashboard += "\n"
        if top_agents:
            dashboard += "🏆 *فعال‌ترین عامل‌ها:*\n"
            for i, (name, count, commission) in enumerate(top_agents, 1):
                commission_text = f"{commission:,.0f}" if commission else "۰"
                dashboard += f"   {i}. {name} - {count} حواله ({commission_text} افغانی)\n"
            dashboard += "\n"
        if least_active_agents:
            dashboard += "📉 *کم‌کارترین عامل‌های فعال:*\n"
            for i, (name, count) in enumerate(least_active_agents, 1):
                dashboard += f"   {i}. {name} - {count} حواله\n"
            dashboard += "\n"
        if top_earners:
            dashboard += "💰 *بالاترین سودآوری عامل‌ها:*\n"
            for i, (name, commission, count) in enumerate(top_earners, 1):
                dashboard += f"   {i}. {name} - {commission:,.0f} افغانی ({count} حواله)\n"
            dashboard += "\n"
        dashboard += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
        dashboard += f"📅 به‌روزرسانی: {dt.now().strftime('%Y/%m/%d %H:%M')}"
        keyboard = [
            ["🔄 بروزرسانی داشبورد"],
            ["💰 مدیریت مالی مرکزی"],
            ["📥 دانلود گزارش اکسل", "🔙 بازگشت به منوی ادمین"],
        ]
    else:
        dashboard = "📈 *پرمختللې احصایوي ډشبورډ*\n"
        dashboard += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
        dashboard += "📊 *د سيستم ټول احصایې:*\n"
        dashboard += f"   👥 عامل‌های فعال: {active_agents}\n"
        dashboard += f"   📦 کل حواله‌ها: {total_transactions:,}\n"
        dashboard += f"   💰 مجموع مبالغ: {total_amount:,.0f} افغانی\n"
        dashboard += f"   💸 مجموع کارمزدها: {total_commission:,.0f} افغانی\n\n"
        dashboard += "📅 *د وختونو په پرتله کړنه:*\n"
        dashboard += f"   🗓️ ۷ روز گذشته: {last_7_days_transactions:,} حواله ({last_7_days_amount:,.0f} افغانی)\n"
        dashboard += f"   🗓️ ۳۰ روز گذشته: {last_30_days_transactions:,} حواله ({last_30_days_amount:,.0f} افغانی)\n"
        if last_7_days_transactions > 0:
            avg_per_day = last_7_days_transactions / 7
            dashboard += f"   📊 میانگین روزانه (۷ روز): {avg_per_day:.1f} حواله\n"
        dashboard += "\n"
        if top_agents:
            dashboard += "🏆 *تر فعالو عاملان:*\n"
            for i, (name, count, commission) in enumerate(top_agents, 1):
                commission_text = f"{commission:,.0f}" if commission else "۰"
                dashboard += f"   {i}. {name} - {count} حواله ({commission_text} افغانی)\n"
            dashboard += "\n"
        if least_active_agents:
            dashboard += "📉 *لږ فعالو عاملان (فعال):*\n"
            for i, (name, count) in enumerate(least_active_agents, 1):
                dashboard += f"   {i}. {name} - {count} حواله\n"
            dashboard += "\n"
        if top_earners:
            dashboard += "💰 *تر ګټې لرونکي عاملان:*\n"
            for i, (name, commission, count) in enumerate(top_earners, 1):
                dashboard += f"   {i}. {name} - {commission:,.0f} افغانی ({count} حواله)\n"
            dashboard += "\n"
        dashboard += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
        dashboard += f"📅 نوول: {dt.now().strftime('%Y/%m/%d %H:%M')}"
        keyboard = [
            ["🔄 د ډشبورډ نو کول"],
            ["💰 د مرکزي مالی مدیریت"],
            ["📥 د اکسل راپور ښکته کول", "🔙 د ادمین مینو ته شاته"],
        ]

    await update.message.reply_text(
        dashboard,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


# =======================
# 📥 گزارش اکسل ادمین
# =======================


@require_admin
async def download_admin_excel_report(update, context):
    # ساخت و ارسال فایل اکسل شامل اطلاعات عامل‌ها، تراکنش‌ها و آمار روزانه
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.excel_preparing", lang=lang)
    )
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT a.id, a.name, a.province, a.phone, a.is_active,
               b.balance, b.currency,
               COUNT(t.id) as transaction_count,
               SUM(t.amount) as total_amount,
               SUM(t.commission) as total_commission
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        GROUP BY a.id, b.currency
        ORDER BY a.id
    """)
    agents_data = cur.fetchall()
    
    cur.execute("""
        SELECT DATE(created_at) as date, COUNT(*) as count,
               SUM(amount) as total_amount, SUM(commission) as total_commission
        FROM transactions
        WHERE status != 'cancelled'
        GROUP BY DATE(created_at)
        ORDER BY date DESC
        LIMIT 30
    """)
    daily_stats = cur.fetchall()
    
    cur.execute("""
        SELECT a.province, COUNT(t.id) as transaction_count,
               SUM(t.amount) as total_amount, SUM(t.commission) as total_commission
        FROM agents a
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        WHERE a.is_active = 1
        GROUP BY a.province
        ORDER BY total_amount DESC
    """)
    province_stats = cur.fetchall()
    
    cur.execute(
        """
        SELECT a.id,
               a.name,
               a.province,
               COALESCE(SUM(CASE WHEN b.currency = 'AFN' THEN b.balance END), 0) AS afn_balance
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        WHERE a.is_active = 1
        GROUP BY a.id, a.name, a.province
        """
    )
    all_agents_balances = cur.fetchall()
    
    cur.execute(
        """
        SELECT agent_id, SUM(commission)
        FROM transactions
        WHERE status != 'cancelled'
          AND DATE(created_at) >= DATE('now', '-30 day')
          AND currency = 'AFN'
        GROUP BY agent_id
        """
    )
    commission_rows = cur.fetchall()
    commission_map = {row[0]: float(row[1] or 0) for row in commission_rows}
    
    liquidity_rows = []
    for agent_id, name, province, afn_balance in all_agents_balances:
        obligations, obligations_end_date = _collect_future_obligations(agent_id, 30)
        obligations_afn = sum(amount for _, _, amount, _ in obligations.get("AFN", []))
        commissions_afn = commission_map.get(agent_id, 0.0)
        projected = afn_balance + commissions_afn - obligations_afn
        liquidity_rows.append(
            [
                agent_id,
                name,
                province,
                float(afn_balance or 0),
                float(obligations_afn or 0),
                float(commissions_afn or 0),
                float(projected or 0),
                obligations_end_date.isoformat(),
            ]
        )
    
    conn.close()
    
    # ایجاد فایل اکسل
    try:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # شیت خلاصه سیستم
            summary_data = {
                'بخش': ['کل عامل‌ها', 'عامل‌های فعال', 'کل حواله‌ها', 'مجموع مبلغ', 'مجموع کارمزد'],
                'مقدار': [
                    len(agents_data),
                    len([a for a in agents_data if a[4] == 1]),
                    sum([a[7] if a[7] is not None and isinstance(a[7], (int, float)) else 0 for a in agents_data]),
                    f"{sum([a[8] if a[8] is not None and isinstance(a[8], (int, float)) else 0 for a in agents_data]):,.0f} افغانی",
                    f"{sum([a[9] if a[9] is not None and isinstance(a[9], (int, float)) else 0 for a in agents_data]):,.0f} افغانی"
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='خلاصه سیستم', index=False)
            
            # شیت اطلاعات عامل‌ها
            agents_df = pd.DataFrame(agents_data, columns=[
                'کد عامل', 'نام', 'ولایت', 'تلفن', 'وضعیت', 'موجودی', 'ارز',
                'تعداد حواله', 'مجموع مبلغ', 'مجموع کارمزد'
            ])
            agents_df['وضعیت'] = agents_df['وضعیت'].apply(lambda x: 'فعال' if x == 1 else 'غیرفعال')
            agents_df.to_excel(writer, sheet_name='عامل‌ها', index=False)
            
            # شیت آمار روزانه
            daily_df = pd.DataFrame(daily_stats, columns=[
                'تاریخ', 'تعداد حواله', 'مبلغ کل', 'کارمزد کل'
            ])
            daily_df.to_excel(writer, sheet_name='آمار روزانه', index=False)

            province_rows = []
            for province, tx_count, total_amount, total_commission in province_stats:
                province_rows.append(
                    [
                        province or "نامشخص",
                        int(tx_count or 0),
                        float(total_amount or 0),
                        float(total_commission or 0),
                    ]
                )

            province_df = pd.DataFrame(
                province_rows,
                columns=['ولایت', 'تعداد حواله', 'مبلغ کل', 'کارمزد کل'],
            )
            province_df.to_excel(writer, sheet_name='آمار ولایت‌ها', index=False)
            
            liquidity_df = pd.DataFrame(
                liquidity_rows,
                columns=[
                    'کد عامل',
                    'نام',
                    'ولایت',
                    'موجودی فعلی AFN',
                    'تعهدات ۳۰ روز آینده AFN',
                    'کمیسیون ۳۰ روز گذشته AFN',
                    'پیش‌بینی نقدینگی AFN',
                    'انتهای بازه تعهدات',
                ],
            )
            liquidity_df.to_excel(writer, sheet_name='ریسک نقدینگی', index=False)

            workbook = writer.book

            if 'آمار روزانه' in workbook.sheetnames:
                sheet_daily = workbook['آمار روزانه']
                daily_chart = LineChart()
                daily_chart.title = "روند تعداد حواله‌ها"
                daily_chart.y_axis.title = "تعداد حواله"
                daily_chart.x_axis.title = "تاریخ"
                data = Reference(
                    sheet_daily,
                    min_col=2,
                    max_col=2,
                    min_row=1,
                    max_row=sheet_daily.max_row,
                )
                daily_chart.add_data(data, titles_from_data=True)
                cats = Reference(
                    sheet_daily,
                    min_col=1,
                    min_row=2,
                    max_row=sheet_daily.max_row,
                )
                daily_chart.set_categories(cats)
                sheet_daily.add_chart(daily_chart, "F2")

            if 'آمار ولایت‌ها' in workbook.sheetnames:
                sheet_province = workbook['آمار ولایت‌ها']
                province_chart = BarChart()
                province_chart.title = "مبلغ کل حواله‌ها به تفکیک ولایت"
                province_chart.y_axis.title = "مبلغ کل"
                province_chart.x_axis.title = "ولایت"
                data = Reference(
                    sheet_province,
                    min_col=3,
                    max_col=3,
                    min_row=1,
                    max_row=sheet_province.max_row,
                )
                province_chart.add_data(data, titles_from_data=True)
                cats = Reference(
                    sheet_province,
                    min_col=1,
                    min_row=2,
                    max_row=sheet_province.max_row,
                )
                province_chart.set_categories(cats)
                sheet_province.add_chart(province_chart, "F2")
        
        output.seek(0)
        
        filename = f"گزارش_ادمین_کامل_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await update.message.reply_document(
            document=output,
            filename=filename,
            caption=f"📊 *گزارش کامل ادمین*\n\n"
                    f"📅 تاریخ: {dt.now().strftime('%Y-%m-%d %H:%M')}\n"
                    f"📦 شامل: اطلاعات عامل‌ها، آمار روزانه، آمار ولایت‌ها\n"
                    f"📊 تحلیل کامل عملکرد سیستم",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.exception("Error creating admin excel report")
        await update.message.reply_text(f"❌ خطا در ایجاد گزارش اکسل: {str(e)}")
