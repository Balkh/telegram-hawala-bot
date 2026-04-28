from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ConversationHandler
import pandas as pd
import io
from datetime import datetime as dt, timedelta
import logging
from openpyxl.chart import BarChart, Reference
import calendar

from bot.services.database import (
    get_agent_by_phone,
    bind_agent_telegram_id,
    get_db,
    get_agent_balance,
    check_sufficient_balance,
    get_agent_by_id,
    increase_failed_attempts,
    reset_failed_attempts,
    lock_agent,
    create_agent_expense,
    create_staff_contract,
    get_staff_contracts_for_agent,
    create_fixed_expense_contract,
    get_fixed_expense_contracts_for_agent,
    get_active_agents_with_telegram,
)
from bot.services.security import verify_password, hash_password, validate_agent_password
from bot.services.auth import require_agent, require_any_auth
from bot.services.receipt import generate_receipt_image
from bot.services.localization import _
from bot.services.dates import (
    gregorian_to_jalali_str,
    jalali_to_gregorian_str,
    today_gregorian_str,
    today_jalali_str,
)

logger = logging.getLogger(__name__)


# این فایل شامل جریان‌های اصلی کار عامل است:
# - احراز هویت و ورود عامل
# - ثبت حواله جدید و مدیریت مراحل آن
# - پرداخت حواله و صدور رسید
# - گزارش‌ها، جستجو و مدیریت موجودی/ارزها برای عامل


def get_lang(context):
    return context.user_data.get("lang", "fa")


def _compute_next_pay_date(start_date_str, pay_day_of_month):
    try:
        start_date = dt.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date = dt.now().date()

    today = dt.now().date()

    base_date = today if today >= start_date else start_date

    year = base_date.year
    month = base_date.month

    _, days_in_month = calendar.monthrange(year, month)
    day = min(max(1, int(pay_day_of_month)), days_in_month)

    candidate = dt(year, month, day).date()

    if candidate >= base_date:
        next_date = candidate
    else:
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1
        _, days_in_month = calendar.monthrange(year, month)
        day = min(max(1, int(pay_day_of_month)), days_in_month)
        next_date = dt(year, month, day).date()

    return next_date.strftime("%Y-%m-%d")


def _compute_next_pay_date_interval_days(start_date_str, interval_days):
    try:
        start_date = dt.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date = dt.now().date()

    today = dt.now().date()
    base_date = today if today >= start_date else start_date

    next_date = start_date
    step = max(1, int(interval_days))

    while next_date <= base_date:
        next_date = next_date + timedelta(days=step)

    return next_date.strftime("%Y-%m-%d")


def _get_next_interval_due_on_or_after(start_date_str, interval_days, from_date):
    try:
        start_date = dt.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

    step = max(1, int(interval_days))
    due = start_date + timedelta(days=step)

    while due < from_date:
        due = due + timedelta(days=step)

    return due


async def send_daily_due_reminders(context):
    today_g = today_gregorian_str()
    today_j = today_jalali_str()
    today_date = dt.strptime(today_g, "%Y-%m-%d").date()

    agents = get_active_agents_with_telegram()

    for agent_row in agents:
        agent_id = agent_row["id"]
        chat_id = agent_row["telegram_id"]

        staff_contracts = get_staff_contracts_for_agent(agent_id)
        fixed_contracts = get_fixed_expense_contracts_for_agent(agent_id)

        staff_lines = []
        fixed_lines = []

        for row in staff_contracts:
            if not row["is_active"]:
                continue
            interval_days = row["pay_day_of_month"]
            if not interval_days or interval_days <= 0:
                continue
            next_due = _get_next_interval_due_on_or_after(
                row["start_date"], interval_days, today_date
            )
            if not next_due:
                continue
            if next_due == today_date:
                monthly_salary = float(row["monthly_salary"] or 0)
                if monthly_salary <= 0:
                    continue
                line = _(
                    "agent.daily_due_staff_line",
                    lang="fa",
                    name=row["employee_name"],
                    amount=f"{monthly_salary:,.0f}",
                    currency=row["currency"],
                )
                staff_lines.append(line)

        for row in fixed_contracts:
            if not row["is_active"]:
                continue
            amount = float(row["amount"] or 0)
            if amount <= 0:
                continue
            frequency = row["frequency"]
            category = row["category"]
            start_date = row["start_date"]
            interval_days = row["pay_day_of_month"]

            due_today = False

            if frequency == "daily":
                try:
                    start_date_obj = dt.strptime(start_date, "%Y-%m-%d").date()
                except ValueError:
                    continue
                due_today = today_date >= start_date_obj
            elif frequency == "monthly" and interval_days:
                if category == "electricity":
                    next_due = _get_next_interval_due_on_or_after(
                        start_date, interval_days, today_date
                    )
                    if not next_due:
                        continue
                else:
                    try:
                        start_date_obj = dt.strptime(start_date, "%Y-%m-%d").date()
                    except ValueError:
                        continue
                    year = today_date.year
                    month = today_date.month
                    _, days_in_month = calendar.monthrange(year, month)
                    day = min(max(1, int(interval_days)), days_in_month)
                    candidate = dt(year, month, day).date()
                    if candidate < start_date_obj:
                        due_today = False
                    else:
                        due_today = candidate == today_date
                if category == "electricity":
                    due_today = next_due == today_date

            if not due_today:
                continue

            line = _(
                "agent.daily_due_fixed_line",
                lang="fa",
                name=row["name"],
                amount=f"{amount:,.0f}",
                currency=row["currency"],
            )
            fixed_lines.append(line)

        if not staff_lines and not fixed_lines:
            continue

        parts = []
        parts.append(
            _(
                "agent.daily_due_title",
                lang="fa",
                today=today_j,
            )
        )

        if staff_lines:
            parts.append(_("agent.daily_due_staff_section", lang="fa"))
            parts.extend(staff_lines)

        if fixed_lines:
            parts.append(_("agent.daily_due_fixed_section", lang="fa"))
            parts.extend(fixed_lines)

        parts.append(_("agent.daily_due_footer", lang="fa"))

        text = "\n".join(parts)

        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                parse_mode="Markdown",
            )
        except Exception as e:
            logger.error(f"Failed to send daily due reminder to {chat_id}: {e}")


def _collect_future_obligations(agent_id, horizon_days):
    today = dt.now().date()
    end_date = today + timedelta(days=horizon_days)

    staff_contracts = get_staff_contracts_for_agent(agent_id)
    fixed_contracts = get_fixed_expense_contracts_for_agent(agent_id)

    obligations_by_currency = {}

    for row in staff_contracts:
        if not row["is_active"]:
            continue
        monthly_salary = float(row["monthly_salary"] or 0)
        if monthly_salary <= 0:
            continue
        interval_days = row["pay_day_of_month"]
        if not interval_days or interval_days <= 0:
            continue
        try:
            start_date = dt.strptime(row["start_date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        due = start_date + timedelta(days=int(interval_days))
        while due < today:
            due = due + timedelta(days=int(interval_days))
        if due > end_date:
            continue
        currency = row["currency"]
        obligations_by_currency.setdefault(currency, [])
        obligations_by_currency[currency].append(
            (
                due,
                row["employee_name"],
                monthly_salary,
                "staff",
            )
        )

    for row in fixed_contracts:
        if not row["is_active"]:
            continue
        amount = float(row["amount"] or 0)
        if amount <= 0:
            continue
        frequency = row["frequency"]
        category = row["category"]
        currency = row["currency"]
        try:
            start_date = dt.strptime(row["start_date"], "%Y-%m-%d").date()
        except ValueError:
            continue

        if frequency == "daily":
            current = today if today >= start_date else start_date
            while current <= end_date:
                obligations_by_currency.setdefault(currency, [])
                obligations_by_currency[currency].append(
                    (
                        current,
                        row["name"],
                        amount,
                        "fixed",
                    )
                )
                current = current + timedelta(days=1)
        elif frequency == "monthly":
            interval_days = row["pay_day_of_month"]
            if not interval_days or interval_days <= 0:
                continue
            if category == "electricity":
                due = start_date + timedelta(days=int(interval_days))
                while due < today:
                    due = due + timedelta(days=int(interval_days))
                if due <= end_date:
                    obligations_by_currency.setdefault(currency, [])
                    obligations_by_currency[currency].append(
                        (
                            due,
                            row["name"],
                            amount,
                            "fixed",
                        )
                    )
            else:
                current = today
                while current <= end_date:
                    year = current.year
                    month = current.month
                    _, days_in_month = calendar.monthrange(year, month)
                    day = min(max(1, int(interval_days)), days_in_month)
                    candidate = dt(year, month, day).date()
                    if candidate < start_date:
                        current = current + timedelta(days=1)
                        continue
                    if candidate < today or candidate > end_date:
                        current = current + timedelta(days=1)
                        continue
                    obligations_by_currency.setdefault(currency, [])
                    obligations_by_currency[currency].append(
                        (
                            candidate,
                            row["name"],
                            amount,
                            "fixed",
                        )
                    )
                    current = end_date + timedelta(days=1)

    return obligations_by_currency, end_date

LOGIN_PHONE, LOGIN_PASSWORD = range(2)
(
    SEND_RECEIVER_AGENT,
    SEND_RECEIVER_NAME,
    SEND_RECEIVER_TAZKIRA,
    SEND_AMOUNT,
    SEND_SENDER_NAME,
    SEND_CURRENCY,
    CONFIRM_TRANSACTION,
    EDIT_TRANSACTION_CHOICE,
    EDIT_RECEIVER_INFO,
    EDIT_AMOUNT,
    TRACK_CODE,
    DELETE_CONFIRM,
    PAY_TRANSACTION_CODE,
    PAY_CONFIRM,
    BALANCE_MENU,
    INCREASE_BALANCE_AMOUNT,
    INCREASE_BALANCE_CURRENCY,
    DECREASE_BALANCE_AMOUNT,
    DECREASE_BALANCE_CURRENCY,
    ADD_CURRENCY_TYPE,
    INCREASE_BALANCE_PHOTO,
    SEARCH_TYPE,
    SEARCH_QUERY,
    SEARCH_DATE_RANGE,
    EXPENSE_CATEGORY,
    EXPENSE_CURRENCY,
    EXPENSE_AMOUNT,
    EXPENSE_DESCRIPTION,
    STAFF_NAME,
    STAFF_CURRENCY,
    STAFF_SALARY,
    STAFF_START_DATE,
    STAFF_PAY_DAY,
    FIXED_EXPENSE_CATEGORY,
    FIXED_EXPENSE_CURRENCY,
    FIXED_EXPENSE_AMOUNT,
    FIXED_EXPENSE_START_DATE,
    FIXED_EXPENSE_PAY_DAY,
    AGENT_CHANGE_PASSWORD,
) = range(39)

# =======================
# 🎛 منوی عامل
# =======================


@require_agent
async def agent_menu(update, context):
    lang = get_lang(context)
    agent_id = context.user_data.get("agent_id")
    
    # 🔔 بررسی حواله‌های در انتظار برای این عامل (به عنوان گیرنده)
    pending_msg = ""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT COUNT(*), currency 
            FROM transactions 
            WHERE receiver_agent_id = ? AND status = 'pending'
            GROUP BY currency
            """,
            (agent_id,),
        )
        pending_counts = cur.fetchall()
        conn.close()

        if pending_counts:
            pending_msg = _( "agent.pending_reminder_title", lang=lang) + "\n"
            for count, currency in pending_counts:
                pending_msg += _(
                    "agent.pending_reminder_item",
                    lang=lang,
                    count=count,
                    currency=currency,
                ) + "\n"
            pending_msg += _("agent.pending_reminder_footer", lang=lang)
    except Exception as e:
        logger.error(f"Error checking pending hawalas for menu: {e}")

    keyboard = [
        [
            _("buttons.agent_menu_mine", lang=lang),
            _("buttons.agent_menu_payable", lang=lang),
        ],
        [
            _("buttons.agent_menu_send", lang=lang),
            _("buttons.agent_menu_search_advanced", lang=lang),
            _("buttons.agent_menu_track_code", lang=lang),
        ],
        [_("buttons.agent_menu_balance_report", lang=lang)],
        [_("buttons.agent_menu_change_password", lang=lang)],
        [_("buttons.agent_menu_logout", lang=lang)],
    ]

    await update.message.reply_text(
        f"{pending_msg}{_('agent.menu_title', lang=lang)}",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


# =======================
# 🔐 ورود عامل
# =======================


async def agent_login_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(_("agent.login_phone", lang=lang))
    return LOGIN_PHONE


async def agent_login_phone(update, context):
    lang = get_lang(context)
    phone = update.message.text.strip()
    agent = get_agent_by_phone(phone)

    if not agent:
        await update.message.reply_text(_("agent.login_not_found", lang=lang))
        return ConversationHandler.END

    agent_id, password_hash, telegram_id, is_active = agent

    if not is_active:
        await update.message.reply_text(_("agent.login_inactive", lang=lang))
        return ConversationHandler.END

    if telegram_id and not context.user_data.get("role"):
        await update.message.reply_text(_("agent.login_already_logged_in", lang=lang))
        return ConversationHandler.END

    context.user_data["login_agent_id"] = agent_id
    context.user_data["password_hash"] = password_hash

    await update.message.reply_text(_("agent.login_password", lang=lang))
    return LOGIN_PASSWORD


async def agent_login_password(update, context):
    lang = get_lang(context)
    password = update.message.text
    agent_id = context.user_data["login_agent_id"]
    hashed = context.user_data["password_hash"]

    if not verify_password(password, hashed):
        increase_failed_attempts(agent_id)
        agent = get_agent_by_id(agent_id)
        attempts = agent["failed_attempts"] if agent and "failed_attempts" in agent.keys() else 0
        if attempts >= 5:
            lock_agent(agent_id)
            await update.message.reply_text(_("agent.login_too_many_attempts", lang=lang))
            return ConversationHandler.END
        await update.message.reply_text(_("agent.login_wrong_password", lang=lang))
        return LOGIN_PASSWORD

    reset_failed_attempts(agent_id)
    telegram_id = update.effective_user.id

    agent = get_agent_by_id(agent_id)
    
    if not agent:
        await update.message.reply_text(_("agent.login_agent_not_found", lang=lang))
        return LOGIN_PASSWORD
    
    if not agent["is_active"]:
        await update.message.reply_text(_("agent.login_agent_blocked", lang=lang))
        return ConversationHandler.END

    bind_agent_telegram_id(agent_id, telegram_id)

    # ذخیره اطلاعات عامل در context
    context.user_data["agent_id"] = agent_id
    context.user_data["role"] = "agent"

    await update.message.reply_text(
        _("agent.login_success", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )

    # نمایش منوی عامل
    await agent_menu(update, context)

    return ConversationHandler.END


@require_agent
async def agent_change_password_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("agent.change_password_old", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.agent_back_to_menu", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return AGENT_CHANGE_PASSWORD


@require_agent
async def agent_change_password_process(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("buttons.agent_back_to_menu", lang=lang):
        await agent_menu(update, context)
        return ConversationHandler.END

    step = context.user_data.get("change_pwd_step", "old")

    if step == "old":
        agent_id = context.user_data.get("agent_id")
        agent = get_agent_by_id(agent_id)
        if not agent:
            await update.message.reply_text(_("agent.login_agent_not_found", lang=lang))
            return ConversationHandler.END

        if not verify_password(text, agent["password_hash"]):
            await update.message.reply_text(
                _("agent.change_password_old_incorrect", lang=lang)
            )
            return AGENT_CHANGE_PASSWORD

        context.user_data["change_pwd_step"] = "new"
        await update.message.reply_text(
            _("agent.change_password_new", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.agent_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return AGENT_CHANGE_PASSWORD

    if step == "new":
        if not validate_agent_password(text):
            await update.message.reply_text(
                _("agent.change_password_new_too_short", lang=lang)
            )
            return AGENT_CHANGE_PASSWORD

        context.user_data["change_pwd_new"] = text
        context.user_data["change_pwd_step"] = "confirm"

        await update.message.reply_text(
            _("agent.change_password_confirm", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.agent_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return AGENT_CHANGE_PASSWORD

    if step == "confirm":
        new_password = context.user_data.get("change_pwd_new")
        if text != new_password:
            await update.message.reply_text(
                _("agent.change_password_not_match", lang=lang)
            )
            context.user_data["change_pwd_step"] = "new"
            await update.message.reply_text(
                _("agent.change_password_new", lang=lang),
                reply_markup=ReplyKeyboardMarkup(
                    [[_("buttons.agent_back_to_menu", lang=lang)]],
                    resize_keyboard=True,
                ),
            )
            return AGENT_CHANGE_PASSWORD

        agent_id = context.user_data.get("agent_id")
        if not agent_id:
            await update.message.reply_text(_("agent.login_agent_not_found", lang=lang))
            return ConversationHandler.END

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "UPDATE agents SET password_hash = ? WHERE id = ?",
            (hash_password(new_password), agent_id),
        )
        conn.commit()
        conn.close()

        context.user_data.pop("change_pwd_step", None)
        context.user_data.pop("change_pwd_new", None)

        await update.message.reply_text(
            _("agent.change_password_success", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.agent_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return ConversationHandler.END

    await agent_menu(update, context)
    return ConversationHandler.END


# =======================
# 💸 ارسال حواله جدید
# =======================


@require_agent
async def send_hawala_start(update, context):
    # شروع جریان چندمرحله‌ای ثبت حواله جدید برای عامل فعلی
    lang = get_lang(context)
    for key in [
        "receiver_agent_id",
        "receiver_name",
        "receiver_tazkira",
        "amount",
        "sender_name",
        "currency",
        "commission",
    ]:
        context.user_data.pop(key, None)

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT id, name, province 
        FROM agents 
        WHERE is_active = 1 AND id != ?
        ORDER BY province
    """,
        (context.user_data["agent_id"],),
    )

    agents = cur.fetchall()
    conn.close()

    if not agents:
        keyboard = [[_("buttons.agent_back_to_menu", lang=lang)]]
        await update.message.reply_text(
            f"{_('agent.no_other_agents_title', lang=lang)}\n\n"
            f"{_('agent.no_other_agents_body', lang=lang)}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return ConversationHandler.END

    text = _("agent.select_receiver_agent_title", lang=lang) + "\n\n"

    # گروه‌بندی بر اساس ولایت
    provinces = {}
    for agent_id, name, province in agents:
        if province not in provinces:
            provinces[province] = []
        provinces[province].append((agent_id, name))

    for province, agent_list in provinces.items():
        text += f"🏙️ *{province}:*\n"
        for agent_id, name in agent_list:
            text += f"   👤 {name} - کد: `{agent_id}`\n"
        text += "\n"

    text += _("agent.enter_receiver_agent_code", lang=lang) + "\n"
    text += _("agent.cancel_hint", lang=lang)

    await update.message.reply_text(
        text, parse_mode="Markdown", reply_markup=ReplyKeyboardRemove()
    )

    return SEND_RECEIVER_AGENT


async def send_receiver_agent(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text in ["/cancel", "❌ لغو"]:
        await update.message.reply_text(
            _("agent.operation_cancelled", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    try:
        receiver_agent_id = int(text)

        conn = get_db()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT id, name, province, is_active 
            FROM agents 
            WHERE id = ? AND is_active = 1 AND id != ?
        """,
            (receiver_agent_id, context.user_data["agent_id"]),
        )

        receiver = cur.fetchone()
        conn.close()

        if not receiver:
            await update.message.reply_text(
                f"{_('agent.receiver_agent_not_found', lang=lang)}\n"
                f"{_('agent.receiver_agent_enter_again', lang=lang)}"
            )
            return SEND_RECEIVER_AGENT

        context.user_data["receiver_agent_id"] = receiver_agent_id
        context.user_data["receiver_agent_name"] = receiver[1]
        context.user_data["receiver_province"] = receiver[2]

        await update.message.reply_text(_("agent.enter_receiver_name", lang=lang))
        return SEND_RECEIVER_NAME

    except ValueError:
        await update.message.reply_text(
            _("agent.receiver_agent_code_not_number", lang=lang)
        )
        return SEND_RECEIVER_AGENT


async def send_receiver_name(update, context):
    lang = get_lang(context)
    context.user_data["receiver_name"] = update.message.text.strip()

    await update.message.reply_text(_("agent.enter_receiver_tazkira", lang=lang))
    return SEND_RECEIVER_TAZKIRA


async def send_receiver_tazkira(update, context):
    lang = get_lang(context)
    tazkira = update.message.text.strip()

    if not tazkira.isdigit():
        await update.message.reply_text(_("agent.tazkira_must_be_number", lang=lang))
        return SEND_RECEIVER_TAZKIRA

    context.user_data["receiver_tazkira"] = tazkira

    await update.message.reply_text(_("agent.enter_amount", lang=lang))
    return SEND_AMOUNT


# قبل از ثبت در دیتابیس، از کاربر نام فرستنده رو بخواهیم:


async def send_amount(update, context):
    lang = get_lang(context)
    try:
        amount = float(update.message.text.strip())

        if amount <= 0:
            await update.message.reply_text(
                _("agent.amount_must_be_positive", lang=lang)
            )
            return SEND_AMOUNT

        context.user_data["amount"] = amount

        keyboard = [["🇦🇫 AFN", "🇺🇸 USD"]]

        await update.message.reply_text(
            _("agent.choose_currency", lang=lang),
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return SEND_CURRENCY

    except ValueError:
        await update.message.reply_text(
            _("agent.amount_invalid_number", lang=lang)
        )
        return SEND_AMOUNT


async def send_sender_name(update, context):
    lang = get_lang(context)
    sender_name = update.message.text.strip()

    if not sender_name or len(sender_name) < 2:
        await update.message.reply_text(
            _("agent.sender_name_too_short", lang=lang)
        )
        return SEND_SENDER_NAME

    context.user_data["sender_name"] = sender_name

    amount = context.user_data["amount"]
    commission = amount * 0.01
    context.user_data["commission"] = commission

    currency = context.user_data["currency"]
    payable = amount - commission

    summary = (
        _("agent.summary_title", lang=lang)
        + "\n\n"
        + _("agent.summary_receiver_agent_label", lang=lang)
        + f" {context.user_data['receiver_agent_name']} ({context.user_data['receiver_province']})\n"
        + _("agent.summary_sender_label", lang=lang)
        + f" {context.user_data['sender_name']}\n"
        + _("agent.summary_receiver_label", lang=lang)
        + f" {context.user_data['receiver_name']}\n"
        + _("agent.summary_tazkira_label", lang=lang)
        + f" {context.user_data['receiver_tazkira']}\n"
        + _("agent.summary_amount_label", lang=lang)
        + f" {amount:,.0f} {currency}\n"
        + _("agent.summary_commission_label", lang=lang)
        + f" {commission:,.0f} {currency}\n"
        + _("agent.summary_payable_label", lang=lang)
        + f" {payable:,.0f} {currency}\n\n"
        + _("agent.summary_confirm_question", lang=lang)
    )

    keyboard = [["✅ تأیید و ثبت", "❌ لغو"]]

    await update.message.reply_text(
        summary,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return CONFIRM_TRANSACTION


async def send_currency(update, context):
    lang = get_lang(context)
    currency_text = update.message.text.strip()

    if "AFN" in currency_text:
        currency = "AFN"
    elif "USD" in currency_text:
        currency = "USD"
    else:
        await update.message.reply_text(
            _("agent.use_buttons_only", lang=lang)
        )
        return SEND_CURRENCY

    context.user_data["currency"] = currency

    await update.message.reply_text(_("agent.enter_sender_name", lang=lang))
    return SEND_SENDER_NAME


@require_agent
async def confirm_transaction(update, context):
    choice = update.message.text.strip()
    lang = get_lang(context)

    if choice == "❌ لغو":
        preserved = {
            "lang": context.user_data.get("lang"),
            "agent_id": context.user_data.get("agent_id"),
            "role": context.user_data.get("role"),
        }
        context.user_data.clear()
        context.user_data.update({k: v for k, v in preserved.items() if v is not None})
        await update.message.reply_text(
            _("agent.operation_cancelled", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    if choice != "✅ تأیید و ثبت":
        await update.message.reply_text(_("agent.use_buttons_only", lang=lang))
        return CONFIRM_TRANSACTION

    try:
        import random

        transaction_code = f"HWL{random.randint(100000, 999999)}"

        agent_id = context.user_data["agent_id"]
        amount = context.user_data["amount"]
        currency = context.user_data["currency"]

        # ثبت در دیتابیس
        conn = get_db()
        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO transactions 
            (transaction_code, agent_id, receiver_agent_id, sender_name, 
             receiver_name, receiver_tazkira, amount, currency, commission, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                transaction_code,
                agent_id,
                context.user_data["receiver_agent_id"],
                context.user_data["sender_name"],
                context.user_data["receiver_name"],
                context.user_data["receiver_tazkira"],
                amount,
                currency,
                context.user_data["commission"],
                "pending",
            ),
        )

        # 🔴 افزایش موجودی عامل مبدأ (چون پول نقد گرفته)
        # مطمئن شو رکورد موجودی وجود داره
        cur.execute(
            """
            SELECT id FROM balances 
            WHERE agent_id = ? AND currency = ?
        """,
            (agent_id, currency),
        )
        if not cur.fetchone():
            cur.execute(
                """
                INSERT INTO balances (agent_id, currency, balance)
                VALUES (?, ?, 0)
            """,
                (agent_id, currency),
            )

        cur.execute(
            """
            UPDATE balances 
            SET balance = balance + ?
            WHERE agent_id = ? AND currency = ?
        """,
            (amount, agent_id, currency),
        )

        conn.commit()

        # 🔔 اطلاع‌رسانی به عامل مقصد
        try:
            receiver_agent_id = context.user_data.get("receiver_agent_id")
            if not receiver_agent_id:
                logger.error("No receiver_agent_id found in context.user_data")
            else:
                db_conn = get_db()
                db_cur = db_conn.cursor()
                db_cur.execute("SELECT telegram_id, name FROM agents WHERE id = ?", (receiver_agent_id,))
                receiver_info = db_cur.fetchone()
                
                if receiver_info and receiver_info[0]:
                    receiver_telegram_id = receiver_info[0]
                    
                    db_cur.execute("SELECT name, province FROM agents WHERE id = ?", (agent_id,))
                    sender_agent_info = db_cur.fetchone()
                    sender_agent_name = sender_agent_info[0] if sender_agent_info else "نامشخص"
                    sender_agent_province = sender_agent_info[1] if sender_agent_info else "نامشخص"

                    notification_text = (
                        _("agent.notify_new_tx_title", lang=lang)
                        + "\n"
                        + "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
                        + _(
                            "agent.notify_new_tx_code_label",
                            lang=lang,
                        )
                        + f" `{transaction_code}`\n"
                        + _(
                            "agent.notify_new_tx_sender_label",
                            lang=lang,
                        )
                        + f" {context.user_data['sender_name']}\n"
                        + _(
                            "agent.notify_new_tx_receiver_label",
                            lang=lang,
                        )
                        + f" {context.user_data['receiver_name']}\n"
                        + _(
                            "agent.notify_new_tx_tazkira_label",
                            lang=lang,
                        )
                        + f" {context.user_data['receiver_tazkira']}\n"
                        + _(
                            "agent.notify_new_tx_amount_label",
                            lang=lang,
                        )
                        + f" {amount:,.0f} {currency}\n"
                        + _(
                            "agent.notify_new_tx_from_agent_label",
                            lang=lang,
                        )
                        + f" {sender_agent_name} ({sender_agent_province})\n"
                        + _(
                            "agent.notify_new_tx_created_at_label",
                            lang=lang,
                        )
                        + f" {dt.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                        + "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
                        + _("agent.notify_new_tx_footer", lang=lang)
                    )
                    
                    # 🔴 تغییر: استفاده از await مستقیم به جای create_task برای اطمینان از ارسال
                    await context.bot.send_message(
                        chat_id=receiver_telegram_id,
                        text=notification_text,
                        parse_mode="Markdown"
                    )
                    logger.info(f"Notification sent successfully to TG: {receiver_telegram_id}")
                else:
                    logger.warning(f"Target agent {receiver_agent_id} has no telegram_id. No notification sent.")
                
                db_conn.close()
        except Exception as notify_err:
            logger.error(f"Failed to send notification: {notify_err}")

        conn.close()

        # دریافت موجودی جدید
        new_balance = get_agent_balance(agent_id, currency)

        keyboard = [
            [_("buttons.agent_menu_send", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        # تولید و ارسال رسید تصویری
        try:
            # دریافت نام و ولایت عامل‌ها برای رسید
            db_conn = get_db()
            db_cur = db_conn.cursor()
            
            # عامل فرستنده
            db_cur.execute("SELECT name, province FROM agents WHERE id = ?", (agent_id,))
            sender_row = db_cur.fetchone()
            if sender_row:
                sender_agent_name, sender_agent_province = sender_row
                sender_agent_display = (
                    f"{sender_agent_name} ({sender_agent_province})"
                    if sender_agent_province
                    else sender_agent_name
                )
            else:
                sender_agent_display = "نامشخص"
            
            # عامل گیرنده
            db_cur.execute(
                "SELECT name, province FROM agents WHERE id = ?",
                (context.user_data["receiver_agent_id"],),
            )
            receiver_row = db_cur.fetchone()
            if receiver_row:
                receiver_agent_name, receiver_agent_province = receiver_row
                receiver_agent_display = (
                    f"{receiver_agent_name} ({receiver_agent_province})"
                    if receiver_agent_province
                    else receiver_agent_name
                )
            else:
                receiver_agent_display = "نامشخص"
            
            db_conn.close()

            receipt_data = {
                "transaction_code": transaction_code,
                "sender_name": context.user_data["sender_name"],
                "receiver_name": context.user_data["receiver_name"],
                "receiver_tazkira": context.user_data["receiver_tazkira"],
                "amount": amount,
                "currency": currency,
                "sender_agent": sender_agent_display,
                "receiver_agent": receiver_agent_display,
                "created_at": dt.now().strftime("%Y-%m-%d %H:%M"),
            }
            
            receipt_img = generate_receipt_image(receipt_data)
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=receipt_img,
                caption=_(
                    "agent.receipt_caption_created",
                    lang=lang,
                    code=transaction_code,
                ),
                parse_mode="Markdown",
            )
        except Exception as receipt_err:
            logger.error(f"Failed to generate/send receipt image: {receipt_err}")

        debug_info = ""
        # اگر برای تست است، نمایش وضعیت نوتیفیکیشن
        try:
            receiver_agent_id = context.user_data.get("receiver_agent_id")
            db_conn = get_db()
            db_cur = db_conn.cursor()
            db_cur.execute("SELECT telegram_id FROM agents WHERE id = ?", (receiver_agent_id,))
            row = db_cur.fetchone()
            if row and row[0]:
                debug_info = f"\n\n📡 *وضعیت نوتیفیکیشن:* ارسال شد به `{row[0]}`"
            else:
                debug_info = f"\n\n⚠️ *وضعیت نوتیفیکیشن:* ارسال نشد (عامل مقصد با آیدی {receiver_agent_id} تلگرام متصل ندارد)"
            db_conn.close()
        except:
            pass

        await update.message.reply_text(
            _(
                "agent.create_success_title",
                lang=lang,
            )
            + "\n\n"
            + _(
                "agent.create_success_code_label",
                lang=lang,
            )
            + f" `{transaction_code}`\n"
            + _(
                "agent.create_success_sender_label",
                lang=lang,
            )
            + f" {context.user_data['sender_name']}\n"
            + _(
                "agent.create_success_receiver_label",
                lang=lang,
            )
            + f" {context.user_data['receiver_name']}\n"
            + _(
                "agent.create_success_tazkira_label",
                lang=lang,
            )
            + f" {context.user_data['receiver_tazkira']}\n"
            + _(
                "agent.create_success_destination_label",
                lang=lang,
            )
            + f" {context.user_data['receiver_province']}\n"
            + _(
                "agent.create_success_amount_label",
                lang=lang,
            )
            + f" {amount:,.0f} {currency}\n"
            + _(
                "agent.create_success_commission_label",
                lang=lang,
            )
            + f" {context.user_data['commission']:,.0f} {currency}\n"
            + _(
                "agent.create_success_payable_label",
                lang=lang,
            )
            + f" {amount - context.user_data['commission']:,.0f} {currency}\n"
            + _(
                "agent.create_success_new_balance_label",
                lang=lang,
            )
            + f" {new_balance:,.0f} {currency}\n\n"
            + _(
                "agent.create_success_hint_title",
                lang=lang,
            )
            + "\n"
            + _(
                "agent.create_success_hint_body",
                lang=lang,
                code=transaction_code,
            )
            + f"{debug_info}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        preserved = {
            "lang": context.user_data.get("lang"),
            "agent_id": context.user_data.get("agent_id"),
            "role": context.user_data.get("role"),
        }
        context.user_data.clear()
        context.user_data.update({k: v for k, v in preserved.items() if v is not None})
        return ConversationHandler.END

    except Exception as e:
        logger.exception("Error in confirm_transaction")
        await update.message.reply_text(
            "❌ خطا در ثبت حواله",
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.agent_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return ConversationHandler.END


# =======================
# 📋 حواله‌های من
# =======================


@require_agent
async def list_my_transactions(update, context):
    lang = get_lang(context)

    conn = get_db()
    cur = conn.cursor()

    # آمار کلی
    cur.execute(
        """
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending_count,
            SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed_count,
            SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_count,
            SUM(amount) as total_amount
        FROM transactions 
        WHERE agent_id = ?
    """,
        (context.user_data["agent_id"],),
    )

    stats = cur.fetchone()
    total, pending_count, completed_count, cancelled_count, total_amount = stats

    # لیست تراکنش‌ها
    cur.execute(
        """
        SELECT 
            t.transaction_code,
            t.receiver_name,
            t.amount,
            t.currency,
            t.status,
            t.created_at,
            a.name as receiver_agent_name,
            a.province as receiver_province
        FROM transactions t
        LEFT JOIN agents a ON t.receiver_agent_id = a.id
        WHERE t.agent_id = ?
        ORDER BY t.created_at DESC
        LIMIT 20
    """,
        (context.user_data["agent_id"],),
    )

    transactions = cur.fetchall()
    conn.close()

    if not transactions:
        await update.message.reply_text(
            f"{_('agent.my_tx_empty_title', lang=lang)}\n\n"
            f"{_('agent.my_tx_empty_hint', lang=lang)}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [_("buttons.agent_menu_send", lang=lang)],
                    [_("buttons.agent_back_to_menu", lang=lang)],
                ],
                resize_keyboard=True,
            ),
        )
        return

    text = _("agent.my_tx_header", lang=lang) + "\n"
    text += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"

    text += _("agent.my_tx_quick_stats_title", lang=lang) + "\n"
    text += _("agent.my_tx_total", lang=lang, count=total or 0) + "\n"
    text += _("agent.my_tx_pending", lang=lang, count=pending_count or 0) + "\n"
    text += _("agent.my_tx_completed", lang=lang, count=completed_count or 0) + "\n"
    text += _("agent.my_tx_cancelled", lang=lang, count=cancelled_count or 0) + "\n"
    if total_amount:
        text += _(
            "agent.my_tx_total_amount",
            lang=lang,
            amount=f"{total_amount:,.0f}",
        ) + "\n"
    text += "\n⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"

    for i, (
        code,
        name,
        amount,
        currency,
        status,
        created_at,
        agent_name,
        province,
    ) in enumerate(transactions, 1):
        if status == "pending":
            status_emoji = "🟡"
            status_text = _("agent.my_tx_status_pending", lang=lang)
            action_note = _("agent.my_tx_action_editable", lang=lang)
        elif status == "completed":
            status_emoji = "🟢"
            status_text = _("agent.my_tx_status_completed", lang=lang)
            action_note = ""
        elif status == "cancelled":
            status_emoji = "🔴"
            status_text = _("agent.my_tx_status_cancelled", lang=lang)
            action_note = ""
        else:
            status_emoji = "⚪"
            status_text = status
            action_note = ""

        text += f"{status_emoji} **{code}** {action_note}\n"
        text += f"   👤 { _('agent.my_tx_receiver_label', lang=lang) }: {name}\n"
        text += f"   📍 { _('agent.my_tx_destination_label', lang=lang) }: {province} ({agent_name})\n"
        text += f"   💰 { _('agent.my_tx_amount_label', lang=lang) }: {amount:,.0f} {currency}\n"
        text += f"   📊 { _('agent.my_tx_status_label', lang=lang) }: {status_text}\n"
        text += f"   📅 { _('agent.my_tx_date_label', lang=lang) }: {created_at[:16]}\n"

        if i < len(transactions):
            text += "\n⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"

    # دکمه‌های عملیاتی
    keyboard = []

    # اگر حواله در انتظار داره، دکمه مدیریت اضافه کن
    if pending_count and pending_count > 0:
        keyboard.append([_("buttons.my_tx_manage_pending", lang=lang)])

    keyboard.append([_("buttons.my_tx_refresh", lang=lang)])
    keyboard.append([_("buttons.agent_back_to_menu", lang=lang)])

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


# =======================
# 🔍 پیگیری با کد حواله
# =======================


@require_agent
async def track_transaction_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("agent.track_enter_code", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )
    return TRACK_CODE


@require_agent
async def track_transaction_code(update, context):
    lang = get_lang(context)
    code = update.message.text.strip().upper()

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT 
            t.transaction_code,
            t.receiver_name,
            t.receiver_tazkira,
            t.amount,
            t.currency,
            t.status,
            t.created_at,
            t.receiver_agent_id,
            a1.name as sender_agent_name,
            a2.name as receiver_agent_name,
            a2.province as receiver_province
        FROM transactions t
        LEFT JOIN agents a1 ON t.agent_id = a1.id
        LEFT JOIN agents a2 ON t.receiver_agent_id = a2.id
        WHERE t.transaction_code = ?
    """,
        (code,),
    )

    transaction = cur.fetchone()
    conn.close()

    if not transaction:
        await update.message.reply_text(_("agent.track_not_found", lang=lang))
        await agent_menu(update, context)
        return ConversationHandler.END

    # نمایش اطلاعات
    (
        code,
        name,
        receiver_tazkira,
        amount,
        currency,
        status,
        created_at,
        receiver_agent_id,
        sender_agent,
        receiver_agent,
        province,
    ) = transaction

    current_agent_id = context.user_data.get("agent_id")
    is_receiver_agent = current_agent_id == receiver_agent_id

    if status == "pending":
        status_emoji = "🟡"
        status_text = _("agent.track_status_pending", lang=lang)
        if is_receiver_agent:
            action_text = "\n" + _("agent.track_you_are_receiver", lang=lang)
        else:
            action_text = "\n" + _("agent.track_receiver_can_go", lang=lang)
    elif status == "completed":
        status_emoji = "🟢"
        status_text = _("agent.track_status_completed", lang=lang)
        action_text = "\n" + _("agent.track_already_completed", lang=lang)
    else:
        status_emoji = "🔴"
        status_text = _("agent.track_status_cancelled", lang=lang)
        action_text = ""

    text = (
        f"{status_emoji} { _('agent.track_status_title', lang=lang) } {status_text}\n\n"
        f"{_('agent.track_code_label', lang=lang)} `{code}`\n"
        f"{_('agent.track_receiver_label', lang=lang)} {name}\n"
        f"{_('agent.track_amount_label', lang=lang)} {amount:,.0f} {currency}\n"
        f"{_('agent.track_destination_label', lang=lang)} {province} ({receiver_agent})\n"
        f"{_('agent.track_created_at_label', lang=lang)} {created_at}\n"
        f"{action_text}"
    )

    from telegram import InlineKeyboardButton, InlineKeyboardMarkup
    keyboard = []
    keyboard.append(
        [
            InlineKeyboardButton(
                _("agent.track_receipt_button", lang=lang),
                callback_data=f"get_receipt_{code}",
            )
        ]
    )
    if is_receiver_agent and status == "pending":
        keyboard.append(
            [
                InlineKeyboardButton(
                    _("agent.track_pay_button", lang=lang),
                    callback_data=f"pay_fast_{code}",
                )
            ]
        )

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    # برای دکمه‌های ریپلای معمولی
    await agent_menu(update, context)
    return ConversationHandler.END


@require_agent
async def list_payable_transactions(update, context):
    lang = get_lang(context)
    agent_id = context.user_data.get("agent_id")
    
    conn = get_db()
    cur = conn.cursor()
    
    # دریافت حواله‌های در انتظار برای این عامل
    cur.execute(
        """
        SELECT 
            t.transaction_code, 
            t.receiver_name, 
            t.amount, 
            t.currency, 
            t.sender_name,
            t.created_at,
            a.name as sender_agent_name
        FROM transactions t
        JOIN agents a ON t.agent_id = a.id
        WHERE t.receiver_agent_id = ? AND t.status = 'pending'
        ORDER BY t.created_at DESC
    """,
        (agent_id,),
    )
    
    payable_list = cur.fetchall()
    conn.close()
    
    if not payable_list:
        await update.message.reply_text(
            _("agent.payable_empty", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.agent_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return
    
    text = _("agent.payable_header", lang=lang) + "\n"
    text += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
    
    for code, receiver, amount, currency, sender, date, sender_agent in payable_list:
        text += f"{_('agent.payable_code_label', lang=lang)} `{code}`\n"
        text += (
            f"{_('agent.payable_receiver_label', lang=lang)} {receiver}\n"
        )
        text += (
            f"{_('agent.payable_amount_label', lang=lang)} {amount:,.0f} {currency}\n"
        )
        text += f"{_('agent.payable_sender_label', lang=lang)} {sender}\n"
        text += (
            f"{_('agent.payable_sender_agent_label', lang=lang)} {sender_agent}\n"
        )
        text += (
            f"{_('agent.payable_date_label', lang=lang)} {date[:16]}\n"
        )
        text += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
        
    text += _("agent.payable_footer_hint", lang=lang)
    
    keyboard = [
        [_("buttons.agent_menu_track_code", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]
    
    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )


@require_agent
async def pay_transaction_start(update, context):
    lang = get_lang(context)
    choice = update.message.text.strip()

    if choice in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    if choice != "💵 پرداخت به گیرنده":
        await update.message.reply_text(_("agent.use_buttons_only", lang=lang))
        return PAY_TRANSACTION_CODE

    code = context.user_data.get("pay_transaction_code")
    if not code:
        await update.message.reply_text(_("agent.pay_code_missing", lang=lang))
        await agent_menu(update, context)
        return ConversationHandler.END

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT 
            t.transaction_code,
            t.receiver_name,
            t.receiver_tazkira,
            t.amount,
            t.currency,
            t.status,
            t.receiver_agent_id
        FROM transactions t
        WHERE t.transaction_code = ? AND t.status = 'pending'
    """,
        (code,),
    )

    transaction = cur.fetchone()
    conn.close()

    if not transaction:
        await update.message.reply_text(
            _("agent.pay_not_found_or_already_paid", lang=lang)
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    (
        code,
        receiver_name,
        receiver_tazkira,
        amount,
        currency,
        status,
        receiver_agent_id,
    ) = transaction

    current_agent_id = context.user_data.get("agent_id")
    if current_agent_id != receiver_agent_id:
        await update.message.reply_text(
            _("agent.pay_not_receiver_agent", lang=lang)
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    # چک موجودی عامل مقصد (باید پول داشته باشد تا بدهد)
    balance = get_agent_balance(current_agent_id, currency)
    if balance < amount:
        keyboard = [[_("buttons.agent_back_to_menu", lang=lang)]]
        text = (
            _("agent.pay_insufficient_balance_title", lang=lang)
            + "\n\n"
            + _(
                "agent.pay_insufficient_balance_amount",
                lang=lang,
                amount=f"{amount:,.0f}",
                currency=currency,
            )
            + "\n"
            + _(
                "agent.pay_insufficient_balance_current",
                lang=lang,
                balance=f"{balance:,.0f}",
                currency=currency,
            )
            + "\n\n"
            + _("agent.pay_insufficient_balance_hint", lang=lang)
        )
        await update.message.reply_text(
            text,
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return ConversationHandler.END

    # ذخیره اطلاعات در context
    context.user_data["pay_transaction_code"] = code
    context.user_data["pay_amount"] = amount
    context.user_data["pay_currency"] = currency
    context.user_data["pay_receiver_name"] = receiver_name
    context.user_data["pay_receiver_tazkira"] = receiver_tazkira

    text = (
        _("agent.pay_confirm_title", lang=lang)
        + "\n\n"
        + _("agent.pay_confirm_code", lang=lang, code=code)
        + "\n"
        + _("agent.pay_confirm_receiver", lang=lang, name=receiver_name)
        + "\n"
        + _(
            "agent.pay_confirm_tazkira",
            lang=lang,
            tazkira=receiver_tazkira,
        )
        + "\n"
        + _(
            "agent.pay_confirm_amount",
            lang=lang,
            amount=f"{amount:,.0f}",
            currency=currency,
        )
        + "\n"
        + _(
            "agent.pay_confirm_balance",
            lang=lang,
            balance=f"{balance:,.0f}",
            currency=currency,
        )
        + "\n\n"
        + _("agent.pay_confirm_check_tazkira", lang=lang)
        + "\n"
        + _("agent.pay_confirm_question", lang=lang)
    )

    keyboard = [["✅ تأیید پرداخت", "❌ انصراف"]]

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return PAY_CONFIRM


@require_agent
async def pay_transaction_confirm(update, context):
    lang = get_lang(context)
    choice = update.message.text.strip()

    if choice == "❌ انصراف":
        keyboard = [[_("buttons.agent_back_to_menu", lang=lang)]]
        await update.message.reply_text(
            _("agent.pay_cancelled", lang=lang),
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        context.user_data.pop("pay_transaction_code", None)
        context.user_data.pop("pay_amount", None)
        context.user_data.pop("pay_currency", None)
        context.user_data.pop("pay_receiver_name", None)
        context.user_data.pop("pay_receiver_tazkira", None)
        return ConversationHandler.END

    if choice != "✅ تأیید پرداخت":
        await update.message.reply_text(_("agent.use_buttons_only", lang=lang))
        return PAY_CONFIRM

    code = context.user_data.get("pay_transaction_code")
    amount = context.user_data.get("pay_amount")
    currency = context.user_data.get("pay_currency")
    receiver_agent_id = context.user_data.get("agent_id")

    if not code or not amount:
        await update.message.reply_text(_("agent.pay_data_missing", lang=lang))
        await agent_menu(update, context)
        return ConversationHandler.END

    conn = get_db()
    cur = conn.cursor()

    try:
        # چک مجدد که حواله هنوز pending است و این عامل مقصد است
        cur.execute(
            """
            SELECT receiver_agent_id, status
            FROM transactions
            WHERE transaction_code = ? AND status = 'pending'
        """,
            (code,),
        )

        row = cur.fetchone()
        if not row or row[0] != receiver_agent_id:
            await update.message.reply_text(
                _("agent.pay_status_changed_or_not_receiver", lang=lang)
            )
            conn.close()
            await agent_menu(update, context)
            return ConversationHandler.END

        # چک موجودی نهایی
        balance = get_agent_balance(receiver_agent_id, currency)
        if balance < amount:
            await update.message.reply_text(
                _("agent.pay_insufficient_balance_short", lang=lang)
                + "\n"
                + _(
                    "agent.pay_insufficient_balance_current_short",
                    lang=lang,
                    balance=f"{balance:,.0f}",
                    currency=currency,
                )
            )
            conn.close()
            await agent_menu(update, context)
            return ConversationHandler.END

        # کسر از موجودی عامل مقصد (چون پول را به گیرنده می‌دهد)
        cur.execute(
            """
            UPDATE balances
            SET balance = balance - ?
            WHERE agent_id = ? AND currency = ?
        """,
            (amount, receiver_agent_id, currency),
        )

        # تغییر وضعیت حواله به completed
        cur.execute(
            """
            UPDATE transactions
            SET status = 'completed', completed_at = CURRENT_TIMESTAMP
            WHERE transaction_code = ? AND status = 'pending'
        """,
            (code,),
        )

        conn.commit()
        conn.close()

        new_balance = get_agent_balance(receiver_agent_id, currency)

        keyboard = [
            [_("buttons.agent_menu_send", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        # تولید و ارسال رسید تصویری پرداخت
        try:
            # دریافت اطلاعات کامل حواله برای رسید
            db_conn = get_db()
            db_cur = db_conn.cursor()
            db_cur.execute(
                """
                SELECT 
                    t.sender_name, 
                    t.receiver_tazkira, 
                    a_sender.name as sender_agent_name, 
                    a_sender.province as sender_agent_province,
                    a_receiver.name as receiver_agent_name,
                    a_receiver.province as receiver_agent_province
                FROM transactions t
                JOIN agents a_sender ON t.agent_id = a_sender.id
                JOIN agents a_receiver ON t.receiver_agent_id = a_receiver.id
                WHERE t.transaction_code = ?
            """,
                (code,),
            )
            row = db_cur.fetchone()
            db_conn.close()

            if row:
                (
                    sender_name,
                    receiver_tazkira,
                    sender_agent_name,
                    sender_agent_province,
                    receiver_agent_name,
                    receiver_agent_province,
                ) = row
                
                sender_agent_display = (
                    f"{sender_agent_name} ({sender_agent_province})"
                    if sender_agent_province
                    else sender_agent_name
                )
                receiver_agent_display = (
                    f"{receiver_agent_name} ({receiver_agent_province})"
                    if receiver_agent_province
                    else receiver_agent_name
                )

                receipt_data = {
                    "transaction_code": code,
                    "sender_name": sender_name,
                    "receiver_name": context.user_data.get("pay_receiver_name"),
                    "receiver_tazkira": receiver_tazkira,
                    "amount": amount,
                    "currency": currency,
                    "sender_agent": sender_agent_display,
                    "receiver_agent": receiver_agent_display,
                    "created_at": dt.now().strftime("%Y-%m-%d %H:%M"),
                }
                
                receipt_img = generate_receipt_image(receipt_data)
                caption = _(
                    "agent.pay_receipt_caption", lang=lang, code=code
                )
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=receipt_img,
                    caption=caption,
                    parse_mode="Markdown",
                )
        except Exception as receipt_err:
            logger.error(f"Failed to generate/send payment receipt image: {receipt_err}")

        text = (
            _("agent.pay_success_title", lang=lang)
            + "\n\n"
            + _("agent.pay_success_code", lang=lang, code=code)
            + "\n"
            + _(
                "agent.pay_success_receiver",
                lang=lang,
                name=context.user_data.get("pay_receiver_name"),
            )
            + "\n"
            + _(
                "agent.pay_success_amount",
                lang=lang,
                amount=f"{amount:,.0f}",
                currency=currency,
            )
            + "\n"
            + _(
                "agent.pay_success_new_balance",
                lang=lang,
                balance=f"{new_balance:,.0f}",
                currency=currency,
            )
            + "\n\n"
            + _("agent.pay_success_footer", lang=lang)
        )

        await update.message.reply_text(
            text,
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        # تمیز کردن context
        context.user_data.pop("pay_transaction_code", None)
        context.user_data.pop("pay_amount", None)
        context.user_data.pop("pay_currency", None)
        context.user_data.pop("pay_receiver_name", None)
        context.user_data.pop("pay_receiver_tazkira", None)

        return ConversationHandler.END

    except Exception:
        conn.close()
        logger.exception("Error completing payment")
        await update.message.reply_text(
            _("agent.pay_generic_error", lang=lang)
        )
        return ConversationHandler.END


@require_agent
async def manage_pending_transactions_start(update, context):
    """شروع مدیریت حواله‌های در انتظار"""
    lang = get_lang(context)
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT transaction_code, receiver_name, amount, currency, created_at
        FROM transactions 
        WHERE agent_id = ? AND status = 'pending'
        ORDER BY created_at DESC
        LIMIT 10
    """,
        (context.user_data["agent_id"],),
    )

    pending = cur.fetchall()
    conn.close()

    if not pending:
        keyboard = [[_("buttons.agent_back_to_menu", lang=lang)]]
        await update.message.reply_text(
            "📭 هیچ حواله در حال انتظاری ندارید.",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return ConversationHandler.END

    # ساخت لیست
    text = "✏️ *حواله‌های در انتظار شما*\n"
    text += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"

    for i, (code, name, amount, currency, created_at) in enumerate(pending, 1):
        text += f"📦 `{code}`\n"
        text += f"   👤 گیرنده: {name}\n"
        text += f"   💰 مبلغ: {amount:,.0f} {currency}\n"
        text += f"   📅 ثبت: {created_at[:16]}\n"

        if i < len(pending):
            text += "   ⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"

    text += "\n⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
    text += "📝 برای ویرایش/حذف، کد حواله را وارد کنید.\n"
    text += "یا از دکمه‌های زیر استفاده کنید:"

    # اضافه کردن دکمه‌ها
    keyboard = [
        [_("buttons.agent_back_to_menu", lang=lang)],
        [_("agent.my_tx_view_all_button", lang=lang)],
    ]

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )
    # شروع Conversation برای انتخاب حواله
    return EDIT_TRANSACTION_CHOICE


@require_agent
async def manage_pending_select_code(update, context):
    """دریافت و بررسی کد حواله در انتظار برای ویرایش/حذف"""
    lang = get_lang(context)
    text = update.message.text.strip().upper()

    # برخورد با دکمه‌های عمومی
    if text in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    if text in [
        "📋 مشاهده همه حواله‌ها",
        _("agent.my_tx_view_all_button", lang=lang),
    ]:
        await list_my_transactions(update, context)
        return ConversationHandler.END

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT transaction_code, receiver_name, amount, currency, created_at
        FROM transactions
        WHERE agent_id = ? AND transaction_code = ? AND status = 'pending'
        """,
        (context.user_data["agent_id"], text),
    )

    row = cur.fetchone()
    conn.close()

    if not row:
        await update.message.reply_text(
            "❌ هیچ حواله در انتظاری با این کد برای شما پیدا نشد.\n"
            "لطفاً کد صحیح را وارد کنید یا با دکمه‌ها بازگردید."
        )
        return EDIT_TRANSACTION_CHOICE

    code, receiver_name, amount, currency, created_at = row

    # ذخیره اطلاعات در context برای مراحل بعدی
    context.user_data["edit_transaction_code"] = code
    context.user_data["edit_transaction_amount"] = amount
    context.user_data["edit_transaction_currency"] = currency

    text = (
        "✏️ *مدیریت حواله انتخاب‌شده*\n\n"
        f"📦 کد: `{code}`\n"
        f"👤 گیرنده: {receiver_name}\n"
        f"💰 مبلغ فعلی: {amount:,.0f} {currency}\n"
        f"📅 ثبت: {created_at[:16]}\n\n"
        "چه کاری می‌خواهید انجام دهید؟"
    )

    keyboard = [
        ["✏️ ویرایش مبلغ", "🗑 لغو حواله"],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return EDIT_TRANSACTION_CHOICE


@require_agent
async def manage_pending_action(update, context):
    """انتخاب نوع عملیات روی حواله: ویرایش مبلغ یا لغو"""
    choice = update.message.text.strip()
    lang = get_lang(context)

    if choice in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    if choice in [
        "📋 مشاهده همه حواله‌ها",
        _("agent.my_tx_view_all_button", lang=lang),
    ]:
        await list_my_transactions(update, context)
        return ConversationHandler.END

    if "edit_transaction_code" not in context.user_data:
        await update.message.reply_text(
            _("agent.manage_pending_select_code_first", lang=lang)
        )
        return EDIT_TRANSACTION_CHOICE

    if choice == "✏️ ویرایش مبلغ":
        await update.message.reply_text(
            _("agent.manage_pending_enter_new_amount", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return EDIT_AMOUNT

    if choice == "🗑 لغو حواله":
        amount = context.user_data["edit_transaction_amount"]
        currency = context.user_data["edit_transaction_currency"]
        code = context.user_data["edit_transaction_code"]

        text = (
            _("agent.manage_pending_cancel_confirm_title", lang=lang)
            + "\n\n"
            + _("agent.manage_pending_cancel_confirm_code_label", lang=lang)
            + f" `{code}`\n"
            + _("agent.manage_pending_cancel_confirm_amount_label", lang=lang)
            + f" {amount:,.0f} {currency}\n\n"
            + _("agent.manage_pending_cancel_confirm_question", lang=lang)
        )

        keyboard = [["✅ تأیید لغو", "❌ انصراف"]]

        await update.message.reply_text(
            text,
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        return DELETE_CONFIRM

    await update.message.reply_text(_("agent.use_buttons_only", lang=lang))
    return EDIT_TRANSACTION_CHOICE


@require_agent
async def edit_pending_amount(update, context):
    """ویرایش مبلغ حواله در انتظار"""
    text = update.message.text.strip()
    lang = get_lang(context)

    try:
        new_amount = float(text)
        if new_amount <= 0:
            await update.message.reply_text(
                _("agent.amount_must_be_positive", lang=lang)
            )
            return EDIT_AMOUNT
    except ValueError:
        await update.message.reply_text(
            _("agent.amount_invalid_number", lang=lang)
        )
        return EDIT_AMOUNT

    if "edit_transaction_code" not in context.user_data:
        await update.message.reply_text(
            _("agent.manage_pending_no_tx_selected_for_edit", lang=lang)
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    old_amount = context.user_data["edit_transaction_amount"]
    currency = context.user_data["edit_transaction_currency"]
    code = context.user_data["edit_transaction_code"]
    agent_id = context.user_data["agent_id"]

    diff = new_amount - old_amount

    conn = get_db()
    cur = conn.cursor()

    try:
        if diff > 0:
            # اگر مبلغ افزایش یافت، موجودی عامل مبدأ هم باید افزایش یابد (پول بیشتر گرفته)
            cur.execute(
                """
                UPDATE balances
                SET balance = balance + ?
                WHERE agent_id = ? AND currency = ?
                """,
                (diff, agent_id, currency),
            )

        elif diff < 0:
            # اگر مبلغ کاهش یافت، موجودی عامل مبدأ هم باید کاهش یابد (باید پول برگرداند)
            cur.execute(
                """
                UPDATE balances
                SET balance = balance + ?
                WHERE agent_id = ? AND currency = ?
                """,
                (diff, agent_id, currency),
            )

        # محاسبه کارمزد جدید (۱٪)
        new_commission = new_amount * 0.01

        # بروزرسانی رکورد حواله
        cur.execute(
            """
            UPDATE transactions
            SET amount = ?, commission = ?
            WHERE transaction_code = ? AND agent_id = ? AND status = 'pending'
            """,
            (new_amount, new_commission, code, agent_id),
        )

        conn.commit()
        conn.close()

        new_balance = get_agent_balance(agent_id, currency)

        keyboard = [
            [_("buttons.my_tx_manage_pending", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _("agent.manage_pending_update_success_title", lang=lang)
            + "\n\n"
            + _("agent.manage_pending_update_success_code_label", lang=lang)
            + f" `{code}`\n"
            + _(
                "agent.manage_pending_update_success_new_amount_label",
                lang=lang,
            )
            + f" {new_amount:,.0f} {currency}\n"
            + _(
                "agent.manage_pending_update_success_commission_label",
                lang=lang,
            )
            + f" {new_commission:,.0f} {currency}\n"
            + _(
                "agent.manage_pending_update_success_balance_label",
                lang=lang,
            )
            + f" {new_balance:,.0f} {currency}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        # تمیز کردن context مربوط به این عملیات
        context.user_data.pop("edit_transaction_code", None)
        context.user_data.pop("edit_transaction_amount", None)
        context.user_data.pop("edit_transaction_currency", None)

        return ConversationHandler.END
    except Exception:
        conn.close()
        logger.exception("Error updating pending transaction amount")
        await update.message.reply_text(
            _("agent.manage_pending_update_error", lang=lang)
        )
        return ConversationHandler.END


@require_agent
async def delete_pending_confirm(update, context):
    """تأیید نهایی لغو حواله در انتظار"""
    choice = update.message.text.strip()
    lang = get_lang(context)

    if choice == "❌ انصراف":
        keyboard = [
            [_("buttons.my_tx_manage_pending", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]
        await update.message.reply_text(
            _("agent.operation_cancelled", lang=lang),
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        context.user_data.pop("edit_transaction_code", None)
        context.user_data.pop("edit_transaction_amount", None)
        context.user_data.pop("edit_transaction_currency", None)
        return ConversationHandler.END

    if choice != "✅ تأیید لغو":
        await update.message.reply_text(_("agent.use_buttons_only", lang=lang))
        return DELETE_CONFIRM

    if "edit_transaction_code" not in context.user_data:
        await update.message.reply_text(
            _("agent.manage_pending_no_tx_selected_for_cancel", lang=lang)
        )
        await agent_menu(update, context)
        return ConversationHandler.END

    code = context.user_data["edit_transaction_code"]
    amount = context.user_data["edit_transaction_amount"]
    currency = context.user_data["edit_transaction_currency"]
    agent_id = context.user_data["agent_id"]

    conn = get_db()
    cur = conn.cursor()

    try:
        # کسر مبلغ از موجودی عامل مبدأ (چون حواله لغو شد و باید پول برگرداند)
        cur.execute(
            """
            UPDATE balances
            SET balance = balance - ?
            WHERE agent_id = ? AND currency = ?
            """,
            (amount, agent_id, currency),
        )

        # علامت‌گذاری حواله به‌عنوان لغو شده (فقط اگر هنوز pending است)
        cur.execute(
            """
            UPDATE transactions
            SET status = 'cancelled'
            WHERE transaction_code = ? AND agent_id = ? AND status = 'pending'
            """,
            (code, agent_id),
        )

        conn.commit()
        conn.close()

        new_balance = get_agent_balance(agent_id, currency)

        keyboard = [
            [_("buttons.agent_menu_send", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _("agent.manage_pending_cancel_success_title", lang=lang)
            + "\n\n"
            + _("agent.manage_pending_cancel_success_code_label", lang=lang)
            + f" `{code}`\n"
            + _(
                "agent.manage_pending_cancel_success_amount_label",
                lang=lang,
            )
            + f" {amount:,.0f} {currency}\n"
            + _(
                "agent.manage_pending_cancel_success_balance_label",
                lang=lang,
            )
            + f" {new_balance:,.0f} {currency}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        context.user_data.pop("edit_transaction_code", None)
        context.user_data.pop("edit_transaction_amount", None)
        context.user_data.pop("edit_transaction_currency", None)

        return ConversationHandler.END
    except Exception:
        conn.close()
        logger.exception("Error cancelling pending transaction")
        await update.message.reply_text(
            _("agent.manage_pending_cancel_error", lang=lang)
        )
        return ConversationHandler.END


# =======================
# 💰 موجودی و گزارش
# =======================


@require_agent
async def balance_and_report_menu(update, context):
    """منوی موجودی و گزارش"""
    lang = get_lang(context)

    keyboard = [
        [_("agent.balance_menu_full_report", lang=lang)],
        [_("agent.balance_menu_download_excel", lang=lang)],
        [_("agent.balance_menu_manage_balance", lang=lang)],
        [_("agent.balance_menu_expenses", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.balance_menu_title", lang=lang)
        + "\n\n"
        + _("agent.balance_menu_hint", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_agent
async def agent_expenses_menu(update, context):
    lang = get_lang(context)

    keyboard = [
        [_("agent.expenses_menu_staff_contracts", lang=lang)],
        [_("agent.expenses_menu_fixed_contracts", lang=lang)],
        [_("agent.expenses_menu_report", lang=lang)],
        [_("agent.expenses_menu_future_obligations", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.expenses_menu_title", lang=lang)
        + "\n\n"
        + _("agent.expenses_menu_hint", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_agent
async def show_future_obligations(update, context):
    lang = get_lang(context)
    agent_id = context.user_data.get("agent_id")

    horizon_days = 30
    obligations_by_currency, end_date = _collect_future_obligations(
        agent_id, horizon_days
    )

    end_date_j = gregorian_to_jalali_str(end_date.strftime("%Y-%m-%d"))

    if not obligations_by_currency:
        await update.message.reply_text(
            _("agent.future_obligations_empty", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [
                    [_("agent.expenses_menu_staff_contracts", lang=lang)],
                    [_("agent.expenses_menu_fixed_contracts", lang=lang)],
                    [_("agent.expenses_menu_report", lang=lang)],
                    [_("buttons.agent_back_to_menu", lang=lang)],
                ],
                resize_keyboard=True,
            ),
        )
        return

    text = _(
        "agent.future_obligations_title",
        lang=lang,
        end_date=end_date_j,
    )
    text += "\n" + _("agent.report_divider", lang=lang) + "\n\n"

    for currency in sorted(obligations_by_currency.keys()):
        items = sorted(obligations_by_currency[currency], key=lambda x: x[0])
        total_amount = 0.0

        text += _("agent.expenses_report_currency_header", lang=lang, currency=currency)
        text += "\n"

        staff_items = [it for it in items if it[3] == "staff"]
        fixed_items = [it for it in items if it[3] == "fixed"]

        if staff_items:
            text += _("agent.future_obligations_staff_section", lang=lang) + "\n"
            for due_date, name, amount, _kind in staff_items:
                total_amount += amount
                date_j = gregorian_to_jalali_str(due_date.strftime("%Y-%m-%d"))
                text += _(
                    "agent.future_obligations_line",
                    lang=lang,
                    name=name,
                    amount=f"{amount:,.0f}",
                    currency=currency,
                    date=date_j,
                )
                text += "\n"
            text += "\n"

        if fixed_items:
            text += _("agent.future_obligations_fixed_section", lang=lang) + "\n"
            for due_date, name, amount, _kind in fixed_items:
                total_amount += amount
                date_j = gregorian_to_jalali_str(due_date.strftime("%Y-%m-%d"))
                text += _(
                    "agent.future_obligations_line",
                    lang=lang,
                    name=name,
                    amount=f"{amount:,.0f}",
                    currency=currency,
                    date=date_j,
                )
                text += "\n"
            text += "\n"

        text += _(
            "agent.future_obligations_summary",
            lang=lang,
            amount=f"{total_amount:,.0f}",
            currency=currency,
        )
        text += "\n\n" + _("agent.report_divider", lang=lang) + "\n\n"

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("agent.expenses_menu_staff_contracts", lang=lang)],
                [_("agent.expenses_menu_fixed_contracts", lang=lang)],
                [_("agent.expenses_menu_report", lang=lang)],
                [_("buttons.agent_back_to_menu", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )


@require_agent
async def show_full_report(update, context):
    """نمایش گزارش کامل: موجودی، آمار حواله‌ها، بدهی/طلب، کمیسیون"""
    lang = get_lang(context)
    agent_id = context.user_data["agent_id"]

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT currency, SUM(balance)
        FROM balances
        WHERE agent_id = ?
        GROUP BY currency
        ORDER BY currency
    """,
        (agent_id,),
    )
    balances = cur.fetchall()

    cur.execute(
        """
        SELECT currency, SUM(commission)
        FROM transactions
        WHERE agent_id = ? AND status != 'cancelled'
        GROUP BY currency
    """,
        (agent_id,),
    )
    commissions = {row[0]: float(row[1] or 0) for row in cur.fetchall()}

    now = dt.now()
    start_30 = now - timedelta(days=30)
    start_30_str = start_30.strftime("%Y-%m-%d %H:%M:%S")
    end_30_str = now.strftime("%Y-%m-%d %H:%M:%S")

    cur.execute(
        """
        SELECT currency, SUM(commission)
        FROM transactions
        WHERE agent_id = ? AND status != 'cancelled'
          AND created_at >= ? AND created_at < ?
        GROUP BY currency
    """,
        (agent_id, start_30_str, end_30_str),
    )
    commissions_last30 = {row[0]: float(row[1] or 0) for row in cur.fetchall()}

    cur.execute(
        """
        SELECT 
            a.name as receiver_name,
            t.currency,
            SUM(t.amount) as debt_amount
        FROM transactions t
        JOIN agents a ON t.receiver_agent_id = a.id
        WHERE t.agent_id = ? AND t.status = 'pending'
        GROUP BY t.receiver_agent_id, t.currency
    """,
        (agent_id,),
    )
    debts = cur.fetchall()

    # ۴. طلب‌های دقیق از عامل‌های دیگر (حواله‌های دریافتی که هنوز پرداخت نشده‌اند)
    cur.execute(
        """
        SELECT 
            a.name as sender_name,
            t.currency,
            SUM(t.amount) as credit_amount
        FROM transactions t
        JOIN agents a ON t.agent_id = a.id
        WHERE t.receiver_agent_id = ? AND t.status = 'pending'
        GROUP BY t.agent_id, t.currency
    """,
        (agent_id,),
    )
    credits = cur.fetchall()

    conn.close()

    staff_contracts = get_staff_contracts_for_agent(agent_id)
    future_obligations_by_currency, _ = _collect_future_obligations(agent_id, 30)

    # ساخت گزارش
    report = _("agent.report_title", lang=lang) + "\n"
    report += _("agent.report_divider", lang=lang) + "\n\n"

    # ۱. درآمد خالص (کمیسیون شما)
    report += _("agent.report_section_commission_title", lang=lang) + "\n"
    report += _("agent.report_section_commission_hint", lang=lang) + "\n"
    if not commissions:
        report += _("agent.report_section_commission_empty", lang=lang) + "\n"
    for curr, comm in commissions.items():
        report += f"✅ {comm:,.0f} {curr}\n"
    report += "\n"

    # ۲. وضعیت بدهی‌ها
    report += _("agent.report_section_debts_title", lang=lang) + "\n"
    report += _("agent.report_section_debts_hint", lang=lang) + "\n"
    if not debts:
        report += _("agent.report_section_debts_empty", lang=lang) + "\n"
    else:
        total_debts = {}
        for name, curr, amount in debts:
            report += f"▪️ {name}: {amount:,.0f} {curr}\n"
            total_debts[curr] = total_debts.get(curr, 0) + amount
        
        report += _("agent.report_divider", lang=lang) + "\n"
        for curr, total in total_debts.items():
            report += _(
                "agent.report_section_debts_total",
                lang=lang,
                amount=f"{total:,.0f}",
                currency=curr,
            ) + "\n"
    report += "\n"

    # ۳. وضعیت طلب‌ها
    report += _("agent.report_section_credits_title", lang=lang) + "\n"
    report += _("agent.report_section_credits_hint", lang=lang) + "\n"
    if not credits:
        report += _("agent.report_section_credits_empty", lang=lang) + "\n"
    else:
        total_credits = {}
        for name, curr, amount in credits:
            report += f"▪️ {name}: {amount:,.0f} {curr}\n"
            total_credits[curr] = total_credits.get(curr, 0) + amount
        
        report += _("agent.report_divider", lang=lang) + "\n"
        for curr, total in total_credits.items():
            report += _(
                "agent.report_section_credits_total",
                lang=lang,
                amount=f"{total:,.0f}",
                currency=curr,
            ) + "\n"
    report += "\n"

    # ۴. تراز نهایی
    report += _("agent.report_section_net_title", lang=lang) + "\n"
    report += _("agent.report_section_net_hint", lang=lang) + "\n"
    
    all_currencies = set(list(commissions.keys()))
    if debts:
        all_currencies.update([d[1] for d in debts])
    if credits:
        all_currencies.update([c[1] for c in credits])
        
    for curr in sorted(all_currencies):
        debt_sum = sum(d[2] for d in debts if d[1] == curr)
        credit_sum = sum(c[2] for c in credits if c[1] == curr)
        net = credit_sum - debt_sum
        emoji = "📈" if net >= 0 else "📉"
        report += f"{emoji} {curr}: {net:,.0f}\n"
    report += "\n"

    # ۵. موجودی فعلی در صندوق
    report += _("agent.report_section_cash_title", lang=lang) + "\n"
    report += _("agent.report_section_cash_hint", lang=lang) + "\n"
    if not balances:
        report += _("agent.report_section_cash_empty", lang=lang) + "\n"
    for curr, bal in balances:
        report += f"💵 {bal:,.0f} {curr}\n"
    
    report += "\n"

    forecast_lines = []
    obligations_sum_by_currency = {}
    for curr, items in future_obligations_by_currency.items():
        total = 0.0
        for due_date, name, amount, kind in items:
            total += float(amount or 0)
        obligations_sum_by_currency[curr] = total

    if balances or obligations_sum_by_currency or commissions_last30:
        report += _("agent.report_section_forecast_title", lang=lang) + "\n"
        report += _("agent.report_section_forecast_hint", lang=lang) + "\n"

        cash_by_currency = {curr: float(bal or 0) for curr, bal in balances}
        all_curr_forecast = set(cash_by_currency.keys()) | set(
            obligations_sum_by_currency.keys()
        ) | set(commissions_last30.keys())

        for curr in sorted(all_curr_forecast):
            cash = cash_by_currency.get(curr, 0.0)
            obligations = obligations_sum_by_currency.get(curr, 0.0)
            comm30 = commissions_last30.get(curr, 0.0)
            net = cash + comm30 - obligations

            report += _(
                "agent.report_section_forecast_line",
                lang=lang,
                currency=curr,
                cash=f"{cash:,.0f}",
                obligations=f"{obligations:,.0f}",
                commission=f"{comm30:,.0f}",
                net=f"{net:,.0f}",
            ) + "\n"

            if net >= 0:
                report += _(
                    "agent.report_section_forecast_positive",
                    lang=lang,
                    currency=curr,
                ) + "\n"
            else:
                report += _(
                    "agent.report_section_forecast_negative",
                    lang=lang,
                    currency=curr,
                ) + "\n"

        report += "\n"

    report += _("agent.report_section_staff_title", lang=lang) + "\n"
    report += _("agent.report_section_staff_hint", lang=lang) + "\n"

    if not staff_contracts:
        report += _("agent.report_section_staff_empty", lang=lang) + "\n"
    else:
        for row in staff_contracts:
            employee_name = row["employee_name"]
            monthly_salary = float(row["monthly_salary"] or 0)
            currency = row["currency"]
            start_date_g = row["start_date"]
            pay_interval_days = row["pay_day_of_month"]
            next_due_g = _compute_next_pay_date_interval_days(start_date_g, pay_interval_days)
            start_date = gregorian_to_jalali_str(start_date_g)
            next_due = gregorian_to_jalali_str(next_due_g)

            report += _(
                "agent.report_section_staff_line",
                lang=lang,
                name=employee_name,
                salary=f"{monthly_salary:,.0f}",
                currency=currency,
                start_date=start_date,
                next_due=next_due,
            ) + "\n"

    report += "\n" + _("agent.report_divider", lang=lang) + "\n"
    report += _("agent.report_footer_summary", lang=lang)

    keyboard = [
        [_("agent.balance_menu_full_report", lang=lang)],
        [_("agent.balance_menu_download_excel", lang=lang)],
        [_("agent.balance_menu_manage_balance", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        report,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )



@require_agent
async def download_excel_report(update, context):
    """تولید و ارسال گزارش اکسل کامل"""
    lang = get_lang(context)
    agent_id = context.user_data["agent_id"]
    
    await update.message.reply_text(_("agent.excel_preparing", lang=lang))
    
    conn = get_db()
    cur = conn.cursor()
    
    # دریافت اطلاعات عامل
    cur.execute("SELECT name, province FROM agents WHERE id = ?", (agent_id,))
    agent_info = cur.fetchone()
    agent_name, agent_province = agent_info if agent_info else ("نامشخص", "نامشخص")
    
    # دریافت تمام حواله‌های عامل
    cur.execute(
        """
        SELECT 
            t.transaction_code,
            t.sender_name,
            t.receiver_name,
            t.receiver_tazkira,
            t.amount,
            t.currency,
            t.commission,
            t.status,
            t.created_at,
            t.completed_at,
            a.name as receiver_agent_name,
            a.province as receiver_province
        FROM transactions t
        LEFT JOIN agents a ON t.receiver_agent_id = a.id
        WHERE t.agent_id = ?
        ORDER BY t.created_at DESC
    """,
        (agent_id,),
    )
    transactions = cur.fetchall()
    
    # دریافت موجودی‌ها
    cur.execute(
        "SELECT currency, SUM(balance) FROM balances WHERE agent_id = ? GROUP BY currency ORDER BY currency",
        (agent_id,),
    )
    balances = cur.fetchall()
    
    conn.close()
    
    # ایجاد DataFrame برای حواله‌ها
    df_transactions = pd.DataFrame(
        transactions,
        columns=[
            _("agent.excel_col_tx_code", lang=lang),
            _("agent.excel_col_sender_name", lang=lang),
            _("agent.excel_col_receiver_name", lang=lang),
            _("agent.excel_col_receiver_tazkira", lang=lang),
            _("agent.excel_col_amount", lang=lang),
            _("agent.excel_col_currency", lang=lang),
            _("agent.excel_col_commission", lang=lang),
            _("agent.excel_col_status", lang=lang),
            _("agent.excel_col_created_at", lang=lang),
            _("agent.excel_col_completed_at", lang=lang),
            _("agent.excel_col_receiver_agent", lang=lang),
            _("agent.excel_col_receiver_province", lang=lang),
        ],
    )
    
    # ایجاد DataFrame برای موجودی‌ها
    df_balances = pd.DataFrame(
        balances,
        columns=[
            _("agent.excel_col_balance_currency", lang=lang),
            _("agent.excel_col_balance_amount", lang=lang),
        ],
    )
    
    # ایجاد فایل اکسل با چند شیت
    try:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # شیت اطلاعات عامل
            agent_data = {
                _("agent.excel_agent_name", lang=lang): [agent_name],
                _("agent.excel_agent_province", lang=lang): [agent_province],
                _(
                    "agent.excel_report_date",
                    lang=lang,
                ): [dt.now().strftime("%Y-%m-%d %H:%M:%S")],
                _("agent.excel_total_transactions", lang=lang): [len(transactions)],
            }
            pd.DataFrame(agent_data).to_excel(
                writer,
                sheet_name=_("agent.excel_sheet_agent_info", lang=lang),
                index=False,
            )
            
            # شیت حواله‌ها
            if not df_transactions.empty:
                df_transactions.to_excel(
                    writer,
                    sheet_name=_("agent.excel_sheet_transactions", lang=lang),
                    index=False,
                )
            
            # شیت موجودی‌ها
            if not df_balances.empty:
                df_balances.to_excel(
                    writer,
                    sheet_name=_("agent.excel_sheet_balances", lang=lang),
                    index=False,
                )
            
            if not df_transactions.empty:
                status_col = _("agent.excel_col_status", lang=lang)
                amount_col = _("agent.excel_col_amount", lang=lang)
                commission_col = _("agent.excel_col_commission", lang=lang)
                created_at_col = _("agent.excel_col_created_at", lang=lang)

                temp_df = df_transactions.copy()
                temp_df["__date"] = temp_df[created_at_col].astype(str).str.slice(0, 10)

                daily_grouped = (
                    temp_df.groupby("__date")
                    .agg(
                        {
                            created_at_col: "count",
                            amount_col: "sum",
                            commission_col: "sum",
                        }
                    )
                    .reset_index()
                    .sort_values("__date")
                )

                daily_date_col = created_at_col
                daily_count_col = _("agent.excel_summary_total_txs", lang=lang)
                daily_amount_col = _("agent.excel_summary_total_amount", lang=lang)
                daily_commission_col = _(
                    "agent.excel_summary_total_commission",
                    lang=lang,
                )

                daily_stats_df = pd.DataFrame(
                    {
                        daily_date_col: daily_grouped["__date"],
                        daily_count_col: daily_grouped[created_at_col],
                        daily_amount_col: daily_grouped[amount_col],
                        daily_commission_col: daily_grouped[commission_col],
                    }
                )

                if lang == "fa":
                    daily_sheet_name = "آمار روزانه"
                else:
                    daily_sheet_name = "ورځنۍ احصایې"

                daily_stats_df.to_excel(
                    writer,
                    sheet_name=daily_sheet_name,
                    index=False,
                )

                summary_type_col = _("agent.excel_summary_type", lang=lang)
                summary_value_col = _("agent.excel_summary_value", lang=lang)

                summary_rows = [
                    [
                        _("agent.excel_summary_total_txs", lang=lang),
                        len(df_transactions),
                    ],
                    [
                        _("agent.excel_summary_pending", lang=lang),
                        len(df_transactions[df_transactions[status_col] == "pending"]),
                    ],
                    [
                        _("agent.excel_summary_completed", lang=lang),
                        len(df_transactions[df_transactions[status_col] == "completed"]),
                    ],
                    [
                        _("agent.excel_summary_cancelled", lang=lang),
                        len(df_transactions[df_transactions[status_col] == "cancelled"]),
                    ],
                    [
                        _("agent.excel_summary_total_amount", lang=lang),
                        float(df_transactions[amount_col].sum() or 0),
                    ],
                    [
                        _("agent.excel_summary_total_commission", lang=lang),
                        float(df_transactions[commission_col].sum() or 0),
                    ],
                ]

                summary_df = pd.DataFrame(
                    summary_rows,
                    columns=[summary_type_col, summary_value_col],
                )

                summary_sheet_name = _("agent.excel_sheet_summary", lang=lang)
                summary_df.to_excel(
                    writer,
                    sheet_name=summary_sheet_name,
                    index=False,
                )

                workbook = writer.book

                if daily_sheet_name in workbook.sheetnames:
                    sheet_daily = workbook[daily_sheet_name]
                    daily_chart = BarChart()
                    if lang == "fa":
                        daily_chart.title = "روند روزانه حواله‌ها"
                    else:
                        daily_chart.title = "د ورځنیو حوالو جریان"
                    daily_chart.y_axis.title = daily_count_col
                    daily_chart.x_axis.title = ""
                    daily_data = Reference(
                        sheet_daily,
                        min_col=2,
                        max_col=2,
                        min_row=1,
                        max_row=sheet_daily.max_row,
                    )
                    daily_chart.add_data(daily_data, titles_from_data=True)
                    daily_cats = Reference(
                        sheet_daily,
                        min_col=1,
                        min_row=2,
                        max_row=sheet_daily.max_row,
                    )
                    daily_chart.set_categories(daily_cats)
                    sheet_daily.add_chart(daily_chart, "F2")

                if summary_sheet_name in workbook.sheetnames:
                    sheet_summary = workbook[summary_sheet_name]
                    chart = BarChart()
                    chart.title = _("agent.excel_summary_chart_title", lang=lang)
                    chart.y_axis.title = summary_value_col
                    chart.x_axis.title = ""
                    data = Reference(
                        sheet_summary,
                        min_col=2,
                        max_col=2,
                        min_row=1,
                        max_row=sheet_summary.max_row,
                    )
                    chart.add_data(data, titles_from_data=True)
                    cats = Reference(
                        sheet_summary,
                        min_col=1,
                        min_row=2,
                        max_row=sheet_summary.max_row,
                    )
                    chart.set_categories(cats)
                    sheet_summary.add_chart(chart, "D2")
        
        output.seek(0)
        
        # ارسال فایل
        filename_prefix = _("agent.excel_filename_prefix", lang=lang)
        filename = f"{filename_prefix}_{agent_name}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await update.message.reply_document(
            document=output,
            filename=filename,
            caption=_("agent.excel_caption_title", lang=lang)
            + "\n\n"
            + _("agent.excel_caption_agent_label", lang=lang)
            + f" {agent_name}\n"
            + _("agent.excel_caption_province_label", lang=lang)
            + f" {agent_province}\n"
            + _("agent.excel_caption_txs_count_label", lang=lang)
            + f" {len(transactions)}\n"
            + _("agent.excel_caption_date_label", lang=lang)
            + f" {dt.now().strftime('%Y-%m-%d %H:%M')}\n\n"
            + _("agent.excel_caption_footer", lang=lang),
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.exception("Error creating agent excel report")
        await update.message.reply_text(
            _(
                "agent.excel_error",
                lang=lang,
                error=str(e),
            )
        )


@require_agent
async def balance_management_menu(update, context):
    """منوی مدیریت موجودی"""
    lang = get_lang(context)
    text = update.message.text.strip() if update.message else ""

    # اگر دکمه بازگشت زده شد
    if text in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return

    keyboard = [
        [_("agent.balance_increase_button", lang=lang)],
        [_("agent.balance_decrease_button", lang=lang)],
        [_("agent.balance_add_currency_button", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.balance_menu_title_short", lang=lang)
        + "\n\n"
        + _("agent.balance_menu_select_option", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_agent
async def add_expense_start(update, context):
    lang = get_lang(context)

    keyboard = [
        [_("agent.expenses_category_staff", lang=lang)],
        [_("agent.expenses_category_food", lang=lang)],
        [_("agent.expenses_category_office", lang=lang)],
        [_("agent.expenses_category_electricity", lang=lang)],
        [_("agent.expenses_category_rent", lang=lang)],
        [_("agent.expenses_category_other", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.expenses_choose_category", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return EXPENSE_CATEGORY


@require_agent
async def add_expense_category(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    mapping = [
        ("staff", "agent.expenses_category_staff"),
        ("food", "agent.expenses_category_food"),
        ("office", "agent.expenses_category_office"),
        ("electricity", "agent.expenses_category_electricity"),
        ("rent", "agent.expenses_category_rent"),
        ("other", "agent.expenses_category_other"),
    ]

    category_key = None
    for key, loc_key in mapping:
        if text == _(loc_key, lang=lang):
            category_key = key
            break

    if not category_key:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return EXPENSE_CATEGORY

    context.user_data["expense_category"] = category_key

    keyboard = [["🇦🇫 AFN", "🇺🇸 USD"], ["🔙 بازگشت"]]

    await update.message.reply_text(
        _("agent.expenses_choose_currency", lang=lang),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return EXPENSE_CURRENCY


@require_agent
async def add_expense_currency(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await agent_expenses_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return EXPENSE_CURRENCY

    context.user_data["expense_currency"] = currency

    await update.message.reply_text(
        _("agent.expenses_enter_amount", lang=lang, currency=currency),
        reply_markup=ReplyKeyboardRemove(),
    )

    return EXPENSE_AMOUNT


@require_agent
async def add_expense_amount(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    try:
        amount = float(text)
        if amount <= 0:
            await update.message.reply_text(
                _("agent.balance_amount_must_be_positive", lang=lang)
            )
            return EXPENSE_AMOUNT
    except ValueError:
        await update.message.reply_text(_("agent.balance_invalid_number", lang=lang))
        return EXPENSE_AMOUNT

    context.user_data["expense_amount"] = amount

    await update.message.reply_text(
        _("agent.expenses_enter_description", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )

    return EXPENSE_DESCRIPTION


@require_agent
async def add_expense_description(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    description = "" if text == "-" else text

    agent_id = context.user_data.get("agent_id")
    category_key = context.user_data.get("expense_category")
    amount = context.user_data.get("expense_amount")
    currency = context.user_data.get("expense_currency")

    if not agent_id or not category_key or amount is None or not currency:
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        return ConversationHandler.END

    create_agent_expense(agent_id, category_key, amount, currency, description)

    category_labels = {
        "staff": _("agent.expenses_category_staff", lang=lang),
        "food": _("agent.expenses_category_food", lang=lang),
        "office": _("agent.expenses_category_office", lang=lang),
        "electricity": _("agent.expenses_category_electricity", lang=lang),
        "rent": _("agent.expenses_category_rent", lang=lang),
        "other": _("agent.expenses_category_other", lang=lang),
    }

    category_label = category_labels.get(category_key, category_key)

    keyboard = [
        [_("agent.expenses_menu_staff_contracts", lang=lang)],
        [_("agent.expenses_menu_fixed_contracts", lang=lang)],
        [_("agent.expenses_menu_report", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _(
            "agent.expenses_saved_success",
            lang=lang,
            category=category_label,
            amount=f"{amount:,.0f}",
            currency=currency,
        ),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    for key in [
        "expense_category",
        "expense_currency",
        "expense_amount",
    ]:
        context.user_data.pop(key, None)

    return ConversationHandler.END


@require_agent
async def staff_contract_start(update, context):
    lang = get_lang(context)

    await update.message.reply_text(
        _("agent.staff_contracts_start_title", lang=lang)
        + "\n\n"
        + _("agent.staff_contracts_enter_name", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )

    return STAFF_NAME


@require_agent
async def staff_contract_name(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if not text:
        await update.message.reply_text(
            _("agent.staff_contracts_enter_name", lang=lang)
        )
        return STAFF_NAME

    context.user_data["staff_employee_name"] = text

    keyboard = [["🇦🇫 AFN", "🇺🇸 USD"], ["🔙 بازگشت"]]

    await update.message.reply_text(
        _("agent.staff_contracts_choose_currency", lang=lang),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return STAFF_CURRENCY


@require_agent
async def staff_contract_currency(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await agent_expenses_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return STAFF_CURRENCY

    context.user_data["staff_currency"] = currency

    await update.message.reply_text(
        _("agent.staff_contracts_enter_salary", lang=lang, currency=currency),
        reply_markup=ReplyKeyboardRemove(),
    )

    return STAFF_SALARY


@require_agent
async def staff_contract_salary(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    try:
        amount = float(text)
        if amount <= 0:
            await update.message.reply_text(
                _("agent.balance_amount_must_be_positive", lang=lang)
            )
            return STAFF_SALARY
    except ValueError:
        await update.message.reply_text(
            _("agent.balance_invalid_number", lang=lang)
        )
        return STAFF_SALARY

    context.user_data["staff_salary"] = amount

    await update.message.reply_text(
        _("agent.staff_contracts_enter_start_date", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("agent.date_pick_today_shamsi", lang=lang)],
                [_("agent.date_pick_manual_shamsi", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )

    return STAFF_START_DATE


@require_agent
async def staff_contract_start_date(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("agent.date_pick_today_shamsi", lang=lang):
        g = today_gregorian_str()
        context.user_data["staff_start_date"] = g
    elif text == _("agent.date_pick_manual_shamsi", lang=lang):
        await update.message.reply_text(
            _("agent.staff_contracts_enter_start_date", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return STAFF_START_DATE
    else:
        try:
            g = jalali_to_gregorian_str(text)
        except ValueError:
            await update.message.reply_text(
                _("agent.staff_contracts_invalid_start_date", lang=lang)
            )
            return STAFF_START_DATE
        context.user_data["staff_start_date"] = g

    await update.message.reply_text(
        _("agent.staff_contracts_enter_pay_day", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )

    return STAFF_PAY_DAY


@require_agent
async def staff_contract_pay_day(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    try:
        pay_day = int(text)
        if pay_day <= 0:
            raise ValueError()
    except ValueError:
        await update.message.reply_text(
            _("agent.staff_contracts_invalid_pay_day", lang=lang)
        )
        return STAFF_PAY_DAY

    agent_id = context.user_data.get("agent_id")
    employee_name = context.user_data.get("staff_employee_name")
    currency = context.user_data.get("staff_currency")
    salary = context.user_data.get("staff_salary")
    start_date_g = context.user_data.get("staff_start_date")

    if (
        not agent_id
        or not employee_name
        or not currency
        or salary is None
        or not start_date_g
    ):
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        return ConversationHandler.END

    create_staff_contract(agent_id, employee_name, salary, currency, start_date_g, pay_day)

    next_due_g = _compute_next_pay_date_interval_days(start_date_g, pay_day)
    start_date_j = gregorian_to_jalali_str(start_date_g)
    next_due_j = gregorian_to_jalali_str(next_due_g)

    keyboard = [
        [_("agent.expenses_menu_staff_contracts", lang=lang)],
        [_("agent.expenses_menu_fixed_contracts", lang=lang)],
        [_("agent.expenses_menu_report", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(_("agent.staff_contracts_saved", lang=lang, name=employee_name, salary=f"{salary:,.0f}", currency=currency, start_date=start_date_j, next_due=next_due_j), parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))

    for key in [
        "staff_employee_name",
        "staff_currency",
        "staff_salary",
        "staff_start_date",
    ]:
        context.user_data.pop(key, None)

    return ConversationHandler.END


@require_agent
async def fixed_expense_start(update, context):
    lang = get_lang(context)

    keyboard = [
        [
            _("agent.fixed_expenses_category_food_daily", lang=lang),
        ],
        [
            _("agent.fixed_expenses_category_electricity_monthly", lang=lang),
        ],
        [
            _("agent.fixed_expenses_category_office_monthly", lang=lang),
        ],
        [
            _("agent.fixed_expenses_category_other_monthly", lang=lang),
        ],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.fixed_expenses_start_title", lang=lang)
        + "\n\n"
        + _("agent.fixed_expenses_choose_category", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return FIXED_EXPENSE_CATEGORY


@require_agent
async def fixed_expense_category(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END

    mapping = [
        (
            "food_daily",
            "agent.fixed_expenses_category_food_daily",
            "food",
            "daily",
        ),
        (
            "electricity_monthly",
            "agent.fixed_expenses_category_electricity_monthly",
            "electricity",
            "monthly",
        ),
        (
            "office_monthly",
            "agent.fixed_expenses_category_office_monthly",
            "office",
            "monthly",
        ),
        (
            "other_monthly",
            "agent.fixed_expenses_category_other_monthly",
            "other",
            "monthly",
        ),
    ]

    selected = None
    for key, loc_key, category, frequency in mapping:
        if text == _(loc_key, lang=lang):
            selected = (key, _(loc_key, lang=lang), category, frequency)
            break

    if not selected:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return FIXED_EXPENSE_CATEGORY

    internal_key, title, category, frequency = selected

    context.user_data["fixed_expense_name"] = title
    context.user_data["fixed_expense_category"] = category
    context.user_data["fixed_expense_frequency"] = frequency

    keyboard = [["🇦🇫 AFN", "🇺🇸 USD"], ["🔙 بازگشت"]]

    await update.message.reply_text(
        _("agent.fixed_expenses_choose_currency", lang=lang),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return FIXED_EXPENSE_CURRENCY


@require_agent
async def fixed_expense_currency(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await agent_expenses_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return FIXED_EXPENSE_CURRENCY

    context.user_data["fixed_expense_currency"] = currency

    await update.message.reply_text(
        _("agent.fixed_expenses_enter_amount", lang=lang, currency=currency),
        reply_markup=ReplyKeyboardRemove(),
    )

    return FIXED_EXPENSE_AMOUNT


@require_agent
async def fixed_expense_amount(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    try:
        amount = float(text)
        if amount <= 0:
            await update.message.reply_text(
                _("agent.balance_amount_must_be_positive", lang=lang)
            )
            return FIXED_EXPENSE_AMOUNT
    except ValueError:
        await update.message.reply_text(_("agent.balance_invalid_number", lang=lang))
        return FIXED_EXPENSE_AMOUNT

    context.user_data["fixed_expense_amount"] = amount

    await update.message.reply_text(
        _("agent.fixed_expenses_enter_start_date", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("agent.date_pick_today_shamsi", lang=lang)],
                [_("agent.date_pick_manual_shamsi", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )

    return FIXED_EXPENSE_START_DATE


@require_agent
async def fixed_expense_start_date(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("agent.date_pick_today_shamsi", lang=lang):
        g = today_gregorian_str()
        context.user_data["fixed_expense_start_date"] = g
    elif text == _("agent.date_pick_manual_shamsi", lang=lang):
        await update.message.reply_text(
            _("agent.fixed_expenses_enter_start_date", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return FIXED_EXPENSE_START_DATE
    else:
        try:
            g = jalali_to_gregorian_str(text)
        except ValueError:
            await update.message.reply_text(
                _("agent.fixed_expenses_invalid_start_date", lang=lang)
            )
            return FIXED_EXPENSE_START_DATE
        context.user_data["fixed_expense_start_date"] = g

    frequency = context.user_data.get("fixed_expense_frequency")

    if frequency == "monthly":
        await update.message.reply_text(
            _("agent.fixed_expenses_enter_pay_day", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return FIXED_EXPENSE_PAY_DAY

    return await fixed_expense_save(update, context, pay_day=None)


@require_agent
async def fixed_expense_pay_day(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    try:
        pay_day = int(text)
        if pay_day < 1 or pay_day > 31:
            raise ValueError()
    except ValueError:
        await update.message.reply_text(
            _("agent.fixed_expenses_invalid_pay_day", lang=lang)
        )
        return FIXED_EXPENSE_PAY_DAY

    return await fixed_expense_save(update, context, pay_day=pay_day)


async def fixed_expense_save(update, context, pay_day):
    lang = get_lang(context)

    agent_id = context.user_data.get("agent_id")
    name = context.user_data.get("fixed_expense_name")
    category = context.user_data.get("fixed_expense_category")
    frequency = context.user_data.get("fixed_expense_frequency")
    currency = context.user_data.get("fixed_expense_currency")
    amount = context.user_data.get("fixed_expense_amount")
    start_date_g = context.user_data.get("fixed_expense_start_date")

    if (
        not agent_id
        or not name
        or not category
        or not frequency
        or not currency
        or amount is None
        or not start_date_g
    ):
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        return ConversationHandler.END

    pay_day_of_month = None
    if frequency == "monthly" and pay_day is not None:
        pay_day_of_month = pay_day

    create_fixed_expense_contract(
        agent_id,
        name,
        category,
        amount,
        currency,
        frequency,
        start_date_g,
        pay_day_of_month,
    )

    frequency_label = "ماهانه" if frequency == "monthly" else "روزانه"

    start_date_j = gregorian_to_jalali_str(start_date_g)

    if frequency == "monthly" and pay_day_of_month:
        if category == "electricity":
            next_due_g = _compute_next_pay_date_interval_days(
                start_date_g, pay_day_of_month
            )
            next_due_j = gregorian_to_jalali_str(next_due_g)
            pay_day_info = (
                "\n📆 موعد پرداخت بعدی (شمسی): "
                + next_due_j
                + f"\n⏱ هر {pay_day_of_month} روز بعد از تاریخ شروع"
            )
        else:
            next_due_g = _compute_next_pay_date(start_date_g, pay_day_of_month)
            next_due_j = gregorian_to_jalali_str(next_due_g)
            pay_day_info = "\n📆 موعد پرداخت بعدی (شمسی): " + next_due_j
    else:
        pay_day_info = ""

    keyboard = [
        [_("agent.expenses_menu_staff_contracts", lang=lang)],
        [_("agent.expenses_menu_fixed_contracts", lang=lang)],
        [_("agent.expenses_menu_report", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.fixed_expenses_saved", lang=lang, name=name, category=category, amount=f"{amount:,.0f}", currency=currency, frequency=frequency_label, start_date=start_date_j, pay_day_info=pay_day_info),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    for key in [
        "fixed_expense_name",
        "fixed_expense_category",
        "fixed_expense_frequency",
        "fixed_expense_currency",
        "fixed_expense_amount",
        "fixed_expense_start_date",
    ]:
        context.user_data.pop(key, None)

    return ConversationHandler.END


@require_agent
async def show_expenses_report(update, context):
    lang = get_lang(context)
    agent_id = context.user_data.get("agent_id")

    text_raw = update.message.text.strip()

    filter_today_label = _("agent.expenses_filter_today", lang=lang)
    filter_7days_label = _("agent.expenses_filter_7days", lang=lang)
    filter_30days_label = _("agent.expenses_filter_30days", lang=lang)
    filter_all_label = _("agent.expenses_filter_all", lang=lang)

    if text_raw in [
        _("agent.expenses_menu_report", lang=lang),
    ]:
        keyboard = [
            [filter_today_label],
            [filter_7days_label],
            [filter_30days_label],
            [filter_all_label],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _("agent.expenses_filter_title", lang=lang),
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return

    filter_type = "all"
    if text_raw == filter_today_label:
        filter_type = "today"
    elif text_raw == filter_7days_label:
        filter_type = "7days"
    elif text_raw == filter_30days_label:
        filter_type = "30days"
    elif text_raw == filter_all_label:
        filter_type = "all"

    now = dt.now()
    start_dt = None
    end_dt = None

    if filter_type == "today":
        start_dt = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = start_dt + timedelta(days=1)
    elif filter_type == "7days":
        end_dt = now
        start_dt = end_dt - timedelta(days=7)
    elif filter_type == "30days":
        end_dt = now
        start_dt = end_dt - timedelta(days=30)

    start_str = None
    end_str = None
    if start_dt and end_dt:
        start_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
        end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()
    cur = conn.cursor()

    commission_sql = """
        SELECT currency, SUM(commission) 
        FROM transactions
        WHERE agent_id = ? AND status != 'cancelled'
    """
    commission_params = [agent_id]

    if start_str and end_str:
        commission_sql += " AND created_at >= ? AND created_at < ?"
        commission_params.extend([start_str, end_str])

    commission_sql += " GROUP BY currency"

    cur.execute(commission_sql, commission_params)
    commission_rows = cur.fetchall()

    expenses_sql = """
        SELECT currency, category, SUM(amount)
        FROM agent_expenses
        WHERE agent_id = ?
    """
    expenses_params = [agent_id]

    if start_str and end_str:
        expenses_sql += " AND created_at >= ? AND created_at < ?"
        expenses_params.extend([start_str, end_str])

    expenses_sql += " GROUP BY currency, category ORDER BY currency"

    cur.execute(expenses_sql, expenses_params)
    expense_rows = cur.fetchall()

    conn.close()

    if not expense_rows:
        await update.message.reply_text(
            _("agent.expenses_report_empty", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [
                    [_("agent.expenses_menu_staff_contracts", lang=lang)],
                    [_("agent.expenses_menu_fixed_contracts", lang=lang)],
                    [_("buttons.agent_back_to_menu", lang=lang)],
                ],
                resize_keyboard=True,
            ),
        )
        return

    commission_by_currency = {}
    for currency, total_commission in commission_rows:
        commission_by_currency[currency] = float(total_commission or 0)

    totals_by_currency = {}
    by_currency_and_category = {}

    for currency, category, total_amount in expense_rows:
        amount = float(total_amount or 0)
        totals_by_currency[currency] = totals_by_currency.get(currency, 0) + amount
        if currency not in by_currency_and_category:
            by_currency_and_category[currency] = {}
        by_currency_and_category[currency][category] = (
            by_currency_and_category[currency].get(category, 0) + amount
        )

    recurring_totals_by_currency = {}
    recurring_lines_by_currency = {}

    if start_dt and end_dt:
        range_start_date = start_dt.date()
        range_end_date = end_dt.date()

        staff_contracts = get_staff_contracts_for_agent(agent_id)
        fixed_contracts = get_fixed_expense_contracts_for_agent(agent_id)

        for row in staff_contracts:
            if not row["is_active"]:
                continue
            monthly_salary = float(row["monthly_salary"] or 0)
            if monthly_salary <= 0:
                continue
            try:
                contract_start = dt.strptime(row["start_date"], "%Y-%m-%d").date()
            except ValueError:
                continue
            active_start = range_start_date if range_start_date > contract_start else contract_start
            active_end = range_end_date
            if active_start >= active_end:
                continue
            active_days = (active_end - active_start).days
            if active_days <= 0:
                continue
            daily_salary = monthly_salary / 30.0
            total_for_period = daily_salary * active_days
            currency = row["currency"]
            recurring_totals_by_currency[currency] = recurring_totals_by_currency.get(currency, 0.0) + total_for_period
            line = _(
                "agent.expenses_report_recurring_line_staff",
                lang=lang,
                name=row["employee_name"],
                amount=f"{total_for_period:,.0f}",
                currency=currency,
                days=active_days,
            )
            if currency not in recurring_lines_by_currency:
                recurring_lines_by_currency[currency] = []
            recurring_lines_by_currency[currency].append(line)

        for row in fixed_contracts:
            if not row["is_active"]:
                continue
            amount = float(row["amount"] or 0)
            if amount <= 0:
                continue
            frequency = row["frequency"]
            if frequency == "daily":
                daily_amount = amount
            elif frequency == "monthly":
                daily_amount = amount / 30.0
            else:
                continue
            try:
                contract_start = dt.strptime(row["start_date"], "%Y-%m-%d").date()
            except ValueError:
                continue
            active_start = range_start_date if range_start_date > contract_start else contract_start
            active_end = range_end_date
            if active_start >= active_end:
                continue
            active_days = (active_end - active_start).days
            if active_days <= 0:
                continue
            total_for_period = daily_amount * active_days
            currency = row["currency"]
            recurring_totals_by_currency[currency] = recurring_totals_by_currency.get(currency, 0.0) + total_for_period
            line = _(
                "agent.expenses_report_recurring_line_fixed",
                lang=lang,
                name=row["name"],
                amount=f"{total_for_period:,.0f}",
                currency=currency,
                days=active_days,
            )
            if currency not in recurring_lines_by_currency:
                recurring_lines_by_currency[currency] = []
            recurring_lines_by_currency[currency].append(line)

    category_labels = {
        "staff": _("agent.expenses_category_staff", lang=lang),
        "food": _("agent.expenses_category_food", lang=lang),
        "office": _("agent.expenses_category_office", lang=lang),
        "electricity": _("agent.expenses_category_electricity", lang=lang),
        "rent": _("agent.expenses_category_rent", lang=lang),
        "other": _("agent.expenses_category_other", lang=lang),
    }

    text = _("agent.expenses_report_title", lang=lang) + "\n"
    text += _("agent.report_divider", lang=lang) + "\n\n"

    net_by_currency = {}

    all_currencies = set(totals_by_currency.keys()) | set(recurring_totals_by_currency.keys()) | set(commission_by_currency.keys())

    days_in_range = None
    if start_dt and end_dt:
        days_in_range = (end_dt.date() - start_dt.date()).days
        if days_in_range <= 0:
            days_in_range = 1

    for currency in sorted(all_currencies):
        total_expenses_manual = totals_by_currency.get(currency, 0.0)
        total_expenses_recurring = recurring_totals_by_currency.get(currency, 0.0)
        total_expenses = total_expenses_manual + total_expenses_recurring
        total_commission = commission_by_currency.get(currency, 0.0)
        net = total_commission - total_expenses

        net_by_currency[currency] = net

        text += _("agent.expenses_report_currency_header", lang=lang, currency=currency)
        text += "\n"

        if total_expenses_manual > 0 and currency in by_currency_and_category:
            text += _("agent.expenses_report_by_category_header", lang=lang) + "\n"
            for category_key, amount in by_currency_and_category[currency].items():
                label = category_labels.get(category_key, category_key)
                text += _(
                    "agent.expenses_report_category_line",
                    lang=lang,
                    category=label,
                    amount=f"{amount:,.0f}",
                    currency=currency,
                )
                text += "\n"
            text += "\n" + _(
                "agent.expenses_report_total",
                lang=lang,
                amount=f"{total_expenses_manual:,.0f}",
                currency=currency,
            )
            text += "\n"

        if total_expenses_recurring > 0 and currency in recurring_lines_by_currency:
            text += _("agent.expenses_report_recurring_header", lang=lang) + "\n"
            for line in recurring_lines_by_currency[currency]:
                text += line + "\n"
            text += _(
                "agent.expenses_report_recurring_total",
                lang=lang,
                amount=f"{total_expenses_recurring:,.0f}",
                currency=currency,
            )
            text += "\n"

        if total_commission > 0 or total_expenses > 0:
            text += "\n" + _(
                "agent.expenses_report_total_commission",
                lang=lang,
                amount=f"{total_commission:,.0f}",
                currency=currency,
            )
            text += "\n" + _(
                "agent.expenses_report_net_profit",
                lang=lang,
                amount=f"{net:,.0f}",
                currency=currency,
            )
            if days_in_range:
                net_daily = net / days_in_range
                text += "\n" + _(
                    "agent.expenses_report_net_daily",
                    lang=lang,
                    amount=f"{net_daily:,.0f}",
                    currency=currency,
                )

            if net > 0:
                text += "\n" + _("agent.expenses_state_positive", lang=lang)
            elif net < 0:
                text += "\n" + _("agent.expenses_state_negative", lang=lang)
            else:
                text += "\n" + _("agent.expenses_state_neutral", lang=lang)

        text += "\n\n" + _("agent.report_divider", lang=lang) + "\n\n"

    if net_by_currency:
        total_net_afn = 0.0
        for currency, net in net_by_currency.items():
            if currency == "AFN":
                total_net_afn += net

        text += _("agent.report_divider", lang=lang) + "\n"
        text += _("agent.expenses_afn_summary_title", lang=lang) + "\n"

        if total_net_afn > 0:
            text += _(
                "agent.expenses_afn_summary_positive",
                lang=lang,
                amount=f"{total_net_afn:,.0f}",
            )
        elif total_net_afn < 0:
            text += _(
                "agent.expenses_afn_summary_negative",
                lang=lang,
                amount=f"{abs(total_net_afn):,.0f}",
            )
        else:
            text += _("agent.expenses_afn_summary_neutral", lang=lang)

        if any(c != "AFN" for c in net_by_currency.keys()):
            text += "\n" + _(
                "agent.expenses_afn_summary_hint_other_currencies",
                lang=lang,
            )

        text += "\n\n"

    fixed_contracts = get_fixed_expense_contracts_for_agent(agent_id)

    if fixed_contracts:
        text += _("agent.report_divider", lang=lang) + "\n"
        text += _("agent.fixed_expenses_start_title", lang=lang) + "\n"

        totals_monthly = {}
        totals_daily = {}

        for row in fixed_contracts:
            amount = float(row["amount"] or 0)
            currency = row["currency"]
            frequency = row["frequency"]
            if frequency == "monthly":
                totals_monthly[currency] = totals_monthly.get(currency, 0) + amount
            elif frequency == "daily":
                totals_daily[currency] = totals_daily.get(currency, 0) + amount

        if totals_monthly:
            text += "\n"
            text += "📆 مجموع مصارف ثابت ماهانه:\n"
            for currency, total in totals_monthly.items():
                text += f"• {total:,.0f} {currency} در ماه\n"

        if totals_daily:
            text += "\n"
            text += "📅 مجموع مصارف ثابت روزانه:\n"
            for currency, total in totals_daily.items():
                text += f"• {total:,.0f} {currency} در روز\n"

        text += "\n"

    keyboard = [
        [_("agent.expenses_menu_staff_contracts", lang=lang)],
        [_("agent.expenses_menu_fixed_contracts", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_agent
async def increase_balance_start(update, context):
    """شروع افزایش موجودی"""
    lang = get_lang(context)
    keyboard = [["🇦🇫 AFN", "🇺🇸 USD"], ["🔙 بازگشت"]]

    await update.message.reply_text(
        _("agent.balance_increase_title", lang=lang)
        + "\n\n"
        + _("agent.balance_increase_choose_currency", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return INCREASE_BALANCE_CURRENCY


@require_agent
async def increase_balance_currency(update, context):
    """انتخاب ارز برای افزایش موجودی"""
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await balance_management_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return INCREASE_BALANCE_CURRENCY

    context.user_data["balance_currency"] = currency
    context.user_data["balance_operation"] = "increase"

    await update.message.reply_text(
        _("agent.balance_enter_amount", lang=lang, currency=currency),
        reply_markup=ReplyKeyboardRemove(),
    )

    return INCREASE_BALANCE_AMOUNT


@require_agent
async def increase_balance_amount(update, context):
    """دریافت مبلغ و افزایش مستقیم موجودی (بدون فیش)"""
    lang = get_lang(context)
    try:
        amount = float(update.message.text.strip())
        if amount <= 0:
            await update.message.reply_text(
                _("agent.balance_amount_must_be_positive", lang=lang)
            )
            return INCREASE_BALANCE_AMOUNT
    except ValueError:
        await update.message.reply_text(
            _("agent.balance_invalid_number", lang=lang)
        )
        return INCREASE_BALANCE_AMOUNT

    agent_id = context.user_data["agent_id"]
    currency = context.user_data["balance_currency"]

    conn = get_db()
    cur = conn.cursor()

    try:
        # اگر قبلاً رکوردی برای این ارز وجود دارد، فقط موجودی را زیاد کن
        cur.execute(
            """
            SELECT id, balance FROM balances
            WHERE agent_id = ? AND currency = ?
        """,
            (agent_id, currency),
        )
        row = cur.fetchone()

        if row:
            cur.execute(
                """
                UPDATE balances
                SET balance = balance + ?
                WHERE agent_id = ? AND currency = ?
            """,
                (amount, agent_id, currency),
            )
        else:
            cur.execute(
                """
                INSERT INTO balances (agent_id, currency, balance)
                VALUES (?, ?, ?)
            """,
                (agent_id, currency, amount),
            )

        conn.commit()
        conn.close()

        new_balance = get_agent_balance(agent_id, currency)

        keyboard = [
            [_("agent.balance_increase_button", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _("agent.balance_increase_success_title", lang=lang)
            + "\n\n"
            + _("agent.balance_increase_success_added_label", lang=lang)
            + f" {amount:,.0f} {currency}\n"
            + _(
                "agent.balance_increase_success_new_balance_label",
                lang=lang,
            )
            + f" {new_balance:,.0f} {currency}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        context.user_data.pop("balance_currency", None)
        context.user_data.pop("balance_operation", None)

        return ConversationHandler.END

    except Exception:
        conn.close()
        logger.exception("Error increasing balance")
        await update.message.reply_text(
            _("agent.balance_increase_error", lang=lang)
        )
        return ConversationHandler.END


@require_agent
async def decrease_balance_start(update, context):
    """شروع کاهش موجودی"""
    lang = get_lang(context)
    keyboard = [["🇦🇫 AFN", "🇺🇸 USD"], ["🔙 بازگشت"]]

    await update.message.reply_text(
        _("agent.balance_decrease_title", lang=lang)
        + "\n\n"
        + _("agent.balance_decrease_choose_currency", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return DECREASE_BALANCE_CURRENCY


@require_agent
async def decrease_balance_currency(update, context):
    """دریافت نوع ارز برای کاهش موجودی"""
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await balance_management_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return DECREASE_BALANCE_CURRENCY

    agent_id = context.user_data["agent_id"]
    balance = get_agent_balance(agent_id, currency)

    if balance <= 0:
        keyboard = [["🔙 بازگشت"]]
        await update.message.reply_text(
            _("agent.balance_zero_balance", lang=lang, currency=currency),
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return DECREASE_BALANCE_CURRENCY

    context.user_data["balance_currency"] = currency
    context.user_data["balance_operation"] = "decrease"

    await update.message.reply_text(
        _(
            "agent.balance_decrease_enter_amount",
            lang=lang,
            currency=currency,
            balance=f"{balance:,.0f}",
        ),
        reply_markup=ReplyKeyboardRemove(),
    )

    return DECREASE_BALANCE_AMOUNT


@require_agent
async def decrease_balance_amount(update, context):
    """دریافت مبلغ و کاهش موجودی"""
    lang = get_lang(context)
    try:
        amount = float(update.message.text.strip())
        if amount <= 0:
            await update.message.reply_text(
                _("agent.balance_amount_must_be_positive", lang=lang)
            )
            return DECREASE_BALANCE_AMOUNT
    except ValueError:
        await update.message.reply_text(
            _("agent.balance_invalid_number", lang=lang)
        )
        return DECREASE_BALANCE_AMOUNT

    agent_id = context.user_data["agent_id"]
    currency = context.user_data["balance_currency"]

    # چک موجودی
    balance = get_agent_balance(agent_id, currency)
    if balance < amount:
        await update.message.reply_text(
            _("agent.balance_not_enough", lang=lang)
            + "\n"
            + _(
                "agent.balance_current_balance_label",
                lang=lang,
            )
            + f" {balance:,.0f} {currency}\n"
            + _(
                "agent.balance_requested_amount_label",
                lang=lang,
            )
            + f" {amount:,.0f} {currency}"
        )
        return DECREASE_BALANCE_AMOUNT

    conn = get_db()
    cur = conn.cursor()

    try:
        # کاهش موجودی
        cur.execute(
            """
            UPDATE balances
            SET balance = balance - ?
            WHERE agent_id = ? AND currency = ?
        """,
            (amount, agent_id, currency),
        )

        conn.commit()
        conn.close()

        new_balance = get_agent_balance(agent_id, currency)

        keyboard = [
            [_("agent.balance_decrease_button", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _("agent.balance_decrease_success_title", lang=lang)
            + "\n\n"
            + _(
                "agent.balance_decrease_success_deducted_label",
                lang=lang,
            )
            + f" {amount:,.0f} {currency}\n"
            + _(
                "agent.balance_decrease_success_new_balance_label",
                lang=lang,
            )
            + f" {new_balance:,.0f} {currency}",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        context.user_data.pop("balance_currency", None)
        context.user_data.pop("balance_operation", None)

        return ConversationHandler.END

    except Exception:
        conn.close()
        logger.exception("Error decreasing balance")
        await update.message.reply_text(
            _("agent.balance_decrease_error", lang=lang)
        )
        return ConversationHandler.END


@require_agent
async def add_currency_start(update, context):
    """شروع اضافه کردن ارز جدید"""
    lang = get_lang(context)
    agent_id = context.user_data["agent_id"]

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT DISTINCT currency
        FROM balances
        WHERE agent_id = ?
    """,
        (agent_id,),
    )
    existing_currencies = [row[0] for row in cur.fetchall()]
    conn.close()

    keyboard = []
    if "AFN" not in existing_currencies:
        keyboard.append(["🇦🇫 AFN"])
    if "USD" not in existing_currencies:
        keyboard.append(["🇺🇸 USD"])

    if not keyboard:
        await update.message.reply_text(
            _("agent.balance_all_currencies_have", lang=lang)
            + "\n"
            + _(
                "agent.balance_your_currencies",
                lang=lang,
                currencies=", ".join(existing_currencies),
            )
        )
        await balance_management_menu(update, context)
        return

    keyboard.append(["🔙 بازگشت"])

    await update.message.reply_text(
        _("agent.balance_add_currency_title", lang=lang)
        + "\n\n"
        + _("agent.balance_add_currency_choose", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )

    return ADD_CURRENCY_TYPE


@require_agent
async def add_currency_confirm(update, context):
    """تأیید و اضافه کردن ارز جدید"""
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == "🔙 بازگشت":
        await balance_management_menu(update, context)
        return ConversationHandler.END

    if "AFN" in text:
        currency = "AFN"
    elif "USD" in text:
        currency = "USD"
    else:
        await update.message.reply_text(_("agent.balance_use_buttons_only", lang=lang))
        return ADD_CURRENCY_TYPE

    agent_id = context.user_data["agent_id"]

    conn = get_db()
    cur = conn.cursor()

    try:
        # چک کن که قبلاً اضافه نشده باشد
        cur.execute(
            """
            SELECT id FROM balances
            WHERE agent_id = ? AND currency = ?
        """,
            (agent_id, currency),
        )
        if cur.fetchone():
            await update.message.reply_text(
                _(
                    "agent.balance_currency_already_exists",
                    lang=lang,
                    currency=currency,
                )
            )
            conn.close()
            await balance_management_menu(update, context)
            return

        # اضافه کردن ارز با موجودی صفر
        cur.execute(
            """
            INSERT INTO balances (agent_id, currency, balance)
            VALUES (?, ?, 0)
        """,
            (agent_id, currency),
        )

        conn.commit()
        conn.close()

        keyboard = [
            [_("agent.balance_increase_button", lang=lang)],
            [_("buttons.agent_back_to_menu", lang=lang)],
        ]

        await update.message.reply_text(
            _(
                "agent.balance_add_currency_success_title",
                lang=lang,
                currency=currency,
            )
            + "\n\n"
            + _(
                "agent.balance_add_currency_current_balance",
                lang=lang,
                currency=currency,
            )
            + "\n"
            + _("agent.balance_add_currency_hint", lang=lang),
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )

        return ConversationHandler.END

    except Exception:
        conn.close()
        logger.exception("Error adding currency")
        await update.message.reply_text(
            _("agent.balance_add_currency_error", lang=lang)
        )
        return ConversationHandler.END


@require_agent
async def search_advanced_start(update, context):
    # ورودی انتخاب نوع جستجوی پیشرفته (نام گیرنده، کد حواله یا امروز)
    lang = get_lang(context)
    keyboard = [
        [_("agent.search_advanced_by_receiver_name", lang=lang)],
        [_("agent.search_advanced_by_code", lang=lang)],
        [_("agent.search_advanced_by_today", lang=lang)],
        [_("buttons.agent_back_to_menu", lang=lang)],
    ]
    await update.message.reply_text(
        _("agent.search_advanced_title", lang=lang)
        + "\n\n"
        + _("agent.search_advanced_hint", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
    return SEARCH_TYPE

@require_agent
async def search_advanced_type(update, context):
    # ذخیره نوع جستجو در context و هدایت به دریافت کوئری/نتایج
    choice = update.message.text.strip()
    lang = get_lang(context)
    
    if choice in [
        "🔙 بازگشت به منوی عامل",
        _("buttons.agent_back_to_menu", lang=lang),
    ]:
        await agent_menu(update, context)
        return ConversationHandler.END
        
    if choice == _("agent.search_advanced_by_receiver_name", lang=lang):
        context.user_data["search_type"] = "by_receiver_name"
        await update.message.reply_text(
            _("agent.search_advanced_enter_receiver_name", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return SEARCH_QUERY
    elif choice == _("agent.search_advanced_by_code", lang=lang):
        context.user_data["search_type"] = "by_code"
        await update.message.reply_text(
            _("agent.search_advanced_enter_code", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        return SEARCH_QUERY
    elif choice == _("agent.search_advanced_by_today", lang=lang):
        context.user_data["search_type"] = "by_today"
        return await search_advanced_results(
            update, context, query=dt.now().strftime("%Y-%m-%d")
        )
    else:
        await update.message.reply_text(
            _("agent.search_advanced_invalid_option", lang=lang)
        )
        return SEARCH_TYPE

@require_agent
async def search_advanced_results(update, context, query=None):
    # اجرای جستجو بر اساس نوع انتخاب‌شده و نمایش حداکثر ۱۰ نتیجهٔ آخر
    lang = get_lang(context)
    if not query:
        query = update.message.text.strip()
    
    search_type = context.user_data.get("search_type", "")
    agent_id = context.user_data.get("agent_id")
    
    conn = get_db()
    cur = conn.cursor()
    
    sql = """
        SELECT t.transaction_code, t.receiver_name, t.amount, t.currency, t.status, t.created_at
        FROM transactions t
        WHERE (t.agent_id = ? OR t.receiver_agent_id = ?)
    """
    params = [agent_id, agent_id]
    
    if search_type == "by_receiver_name":
        sql += " AND t.receiver_name LIKE ?"
        params.append(f"%{query}%")
    elif search_type == "by_code":
        sql += " AND t.transaction_code = ?"
        params.append(query.upper())
    elif search_type == "by_today":
        sql += " AND t.created_at LIKE ?"
        params.append(f"{query}%")
        
    sql += " ORDER BY t.created_at DESC LIMIT 10"
    
    cur.execute(sql, params)
    results = cur.fetchall()
    conn.close()
    
    if not results:
        await update.message.reply_text(
            _("agent.search_advanced_no_results", lang=lang)
        )
        # ماندن در منوی جستجوی پیشرفته برای جستجوی دوباره
        await search_advanced_start(update, context)
        return SEARCH_TYPE
        
    text = _(
        "agent.search_advanced_results_title",
        lang=lang,
        query=query,
    ) + "\n\n"
    keyboard = []
    for code, name, amount, currency, status, created_at in results:
        status_emoji = "🟢" if status == 'completed' else "🟡" if status == 'pending' else "🔴"
        text += f"{status_emoji} `{code}` | {name}\n"
        text += f"💰 {amount:,.0f} {currency} | 📅 {created_at[:16]}\n"
        text += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
        keyboard.append(
            [
                InlineKeyboardButton(
                    _(
                        "agent.search_advanced_receipt_button",
                        lang=lang,
                        code=code,
                    ),
                    callback_data=f"get_receipt_{code}",
                )
            ]
        )
        
    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    # بعد از نمایش نتایج، دوباره منوی جستجوی پیشرفته را نشان بده
    await search_advanced_start(update, context)
    return SEARCH_TYPE

@require_any_auth
async def handle_receipt_callback(update, context):
    """هندلر دکمه شیشه‌ای دانلود رسید"""
    query = update.callback_query
    await query.answer()
    
    code = query.data.replace("get_receipt_", "")
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            t.transaction_code, 
            t.sender_name, 
            t.receiver_name, 
            t.receiver_tazkira, 
            t.amount, 
            t.currency, 
            t.created_at,
            a_sender.name as sender_agent_name, 
            a_sender.province as sender_agent_province,
            a_receiver.name as receiver_agent_name,
            a_receiver.province as receiver_agent_province
        FROM transactions t
        JOIN agents a_sender ON t.agent_id = a_sender.id
        JOIN agents a_receiver ON t.receiver_agent_id = a_receiver.id
        WHERE t.transaction_code = ?
    """, (code,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        lang = get_lang(context)
        await query.message.reply_text(
            _("agent.receipt_not_found", lang=lang)
        )
        return

    sender_agent_name = row[7]
    sender_agent_province = row[8]
    receiver_agent_name = row[9]
    receiver_agent_province = row[10]
    
    sender_agent_display = (
        f"{sender_agent_name} ({sender_agent_province})"
        if sender_agent_province
        else sender_agent_name
    )
    receiver_agent_display = (
        f"{receiver_agent_name} ({receiver_agent_province})"
        if receiver_agent_province
        else receiver_agent_name
    )

    receipt_data = {
        "transaction_code": row[0],
        "sender_name": row[1],
        "receiver_name": row[2],
        "receiver_tazkira": row[3],
        "amount": row[4],
        "currency": row[5],
        "sender_agent": sender_agent_display,
        "receiver_agent": receiver_agent_display,
        "created_at": row[6],
    }
    
    try:
        lang = get_lang(context)
        receipt_img = generate_receipt_image(receipt_data)
        await context.bot.send_photo(
            chat_id=update.effective_chat.id,
            photo=receipt_img,
            caption=_(
                "agent.receipt_caption_duplicate",
                lang=lang,
                code=code,
            ),
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Error sending receipt: {e}")
        await query.message.reply_text(
            _("agent.receipt_error", lang=lang)
        )


@require_agent
async def handle_pay_fast_callback(update, context):
    """هندلر دکمه شیشه‌ای پرداخت سریع"""
    query = update.callback_query
    await query.answer()
    
    code = query.data.replace("pay_fast_", "")
    
    # انتقال به جریان پرداخت
    context.user_data["pay_transaction_code"] = code
    
    # فراخوانی هندلر موجود برای شروع پرداخت
    # توجه: چون در همین فایل هستیم، مستقیم فراخوانی می‌کنیم
    return await pay_transaction_start(update, context)

# =======================
# 🚪 خروج عامل
# =======================


async def agent_logout(update, context):
    """خروج عامل از سیستم"""
    user_id = update.effective_user.id

    # استفاده از تابع unbind
    from bot.services.database import unbind_agent_telegram_id

    unbind_agent_telegram_id(user_id)

    # پاک کردن context
    context.user_data.clear()

    lang = get_lang(context)
    await update.message.reply_text(
        _("agent.logout_success", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )

    # برگشت به منوی اصلی
    from bot.handlers.start import start

    await start(update, context)
