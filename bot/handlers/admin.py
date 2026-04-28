import io
import os
import datetime
import pandas as pd
from datetime import datetime as dt

from telegram import (
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.ext import ConversationHandler
import logging
from openpyxl.chart import BarChart, Reference

from bot.services.errors import global_error_handler
from bot.services.security import (
    hash_password,
    verify_password,
    validate_agent_password,
    validate_admin_password,
)
from bot.services.database import (
    get_db,
    reset_failed_attempts,
    get_admin_by_telegram_id,
    get_all_admins,
)
from bot.services.auth import require_admin
from bot.services.localization import _
from bot.handlers.agent import _collect_future_obligations

logger = logging.getLogger(__name__)


def get_lang(context):
    return context.user_data.get("lang", "fa")

# حالت‌های مکالمه
(
    NAME,
    PASSWORD,
    CONFIRM_PASSWORD,
    PROVINCE,
    PHONE,
    TAZKIRA,
    BALANCE,
    CURRENCY,
    CONFIRM_AGENT,
    TOGGLE_AGENT,
    ADMIN_SEARCH_TX,
    ADMIN_AGENT_EXPENSE,
    AGENT_PASSWORD_RESET,
    ADMIN_CHANGE_PASSWORD,
    ADMIN_CHANGE_USERNAME,
) = range(15)

# =======================
# 👑 د ادمین مینو
# =======================


@require_admin
async def admin_menu(update, context):
    lang = get_lang(context)
    keyboard = [
        [
            _("buttons.admin_menu_create_agent", lang=lang),
            _("buttons.admin_menu_list_agents", lang=lang),
        ],
        [
            _("buttons.admin_menu_search_transactions", lang=lang),
            _("buttons.admin_menu_toggle_agent", lang=lang),
        ],
        [
            _("buttons.admin_menu_search_agents", lang=lang),
            _("buttons.admin_menu_financial_report", lang=lang),
        ],
        [
            _("buttons.admin_menu_download_excel", lang=lang),
            _("buttons.admin_menu_dashboard", lang=lang),
        ],
        [
            _("buttons.admin_menu_agent_expense_overview", lang=lang),
            _("buttons.admin_menu_backup_db", lang=lang),
        ],
        [
            _("buttons.admin_menu_security", lang=lang),
        ],
        [
            _("buttons.admin_menu_logout", lang=lang),
        ],
    ]

    await update.message.reply_text(
        _("admin.menu_title", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_admin
async def admin_agent_expense_overview_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.agent_expense_enter_id", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.back", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return ADMIN_AGENT_EXPENSE


@require_admin
async def admin_agent_expense_overview_show(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("buttons.back", lang=lang):
        await admin_menu(update, context)
        return ConversationHandler.END

    try:
        agent_id = int(text)
    except ValueError:
        await update.message.reply_text(_("admin.agent_expense_not_found", lang=lang))
        return ConversationHandler.END

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT id, name, province
        FROM agents
        WHERE id = ? AND is_active = 1
        """,
        (agent_id,),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        await update.message.reply_text(_("admin.agent_expense_not_found", lang=lang))
        return ConversationHandler.END

    agent_id_db, name, province = row

    cur.execute(
        """
        SELECT COALESCE(SUM(balance), 0)
        FROM balances
        WHERE agent_id = ? AND currency = 'AFN'
        """,
        (agent_id_db,),
    )
    balance_afn = float(cur.fetchone()[0] or 0)

    cur.execute(
        """
        SELECT 
            COUNT(*) as tx_count,
            COALESCE(SUM(amount), 0) as total_amount,
            COALESCE(SUM(commission), 0) as total_commission
        FROM transactions
        WHERE agent_id = ?
          AND status != 'cancelled'
          AND currency = 'AFN'
          AND DATE(created_at) >= DATE('now', '-30 day')
        """,
        (agent_id_db,),
    )
    tx_row = cur.fetchone()
    tx_count_30d = int(tx_row[0] or 0)
    amount_30d = float(tx_row[1] or 0)
    commission_30d = float(tx_row[2] or 0)

    conn.close()

    obligations, _end_date = _collect_future_obligations(agent_id_db, 30)
    obligations_afn = sum(amount for _, _, amount, _ in obligations.get("AFN", []))
    projected = balance_afn + commission_30d - obligations_afn

    title = _(
        "admin.agent_expense_summary_title",
        lang=lang,
        name=name,
        id=agent_id_db,
    )
    body = _(
        "admin.agent_expense_summary_body",
        lang=lang,
        province=province or "-",
        tx_count_30d=tx_count_30d,
        amount_30d=f"{amount_30d:,.0f}",
        commission_30d=f"{commission_30d:,.0f}",
        balance_afn=f"{balance_afn:,.0f}",
        obligations_30d=f"{obligations_afn:,.0f}",
        projected_liquidity=f"{projected:,.0f}",
    )

    await update.message.reply_text(title + "\n" + body, parse_mode="Markdown")
    await admin_menu(update, context)
    return ConversationHandler.END


@require_admin
async def reset_agent_password_start(update, context):
    lang = get_lang(context)
    context.user_data.clear()
    context.user_data["lang"] = lang
    await update.message.reply_text(
        _("admin.reset_agent_password_enter_id", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return AGENT_PASSWORD_RESET


@require_admin
async def reset_agent_password_process(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    step = context.user_data.get("reset_step", "ask_id")

    if step == "ask_id":
        try:
            agent_id = int(text)
        except ValueError:
            await update.message.reply_text(_("admin.toggle_agent_id_must_be_number", lang=lang))
            return AGENT_PASSWORD_RESET

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM agents WHERE id = ?", (agent_id,))
        row = cur.fetchone()
        conn.close()

        if not row:
            await update.message.reply_text(
                _("admin.reset_agent_password_id_not_found", lang=lang)
            )
            return AGENT_PASSWORD_RESET

        context.user_data["reset_agent_id"] = agent_id
        context.user_data["reset_step"] = "new_password"

        await update.message.reply_text(
            _("admin.reset_agent_password_enter_new", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.cancel", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return AGENT_PASSWORD_RESET

    if step == "new_password":
        if not validate_agent_password(text):
            await update.message.reply_text(
                _("admin.reset_agent_password_too_short", lang=lang)
            )
            return AGENT_PASSWORD_RESET

        context.user_data["reset_new_password"] = text
        context.user_data["reset_step"] = "confirm_password"

        await update.message.reply_text(
            _("admin.reset_agent_password_confirm", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.cancel", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return AGENT_PASSWORD_RESET

    if step == "confirm_password":
        new_password = context.user_data.get("reset_new_password")
        if text != new_password:
            await update.message.reply_text(
                _("admin.reset_agent_password_not_match", lang=lang)
            )
            context.user_data["reset_step"] = "new_password"
            await update.message.reply_text(
                _("admin.reset_agent_password_enter_new", lang=lang),
                reply_markup=ReplyKeyboardMarkup(
                    [[_("buttons.cancel", lang=lang)]],
                    resize_keyboard=True,
                ),
            )
            return AGENT_PASSWORD_RESET

        agent_id = context.user_data.get("reset_agent_id")
        if not agent_id:
            await update.message.reply_text(
                _("admin.reset_agent_password_id_not_found", lang=lang)
            )
            await admin_menu(update, context)
            return ConversationHandler.END

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "UPDATE agents SET password_hash = ? WHERE id = ?",
            (hash_password(new_password), agent_id),
        )
        conn.commit()
        conn.close()

        context.user_data.clear()
        context.user_data["lang"] = lang

        await update.message.reply_text(
            _("admin.reset_agent_password_success", lang=lang)
        )
        await admin_menu(update, context)
        return ConversationHandler.END

    await admin_menu(update, context)
    return ConversationHandler.END


@require_admin
async def admin_change_password_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.change_password_old", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.admin_back_to_menu", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    context.user_data["admin_change_step"] = "old"
    return ADMIN_CHANGE_PASSWORD


@require_admin
async def admin_change_password_process(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("buttons.admin_back_to_menu", lang=lang):
        await admin_menu(update, context)
        return ConversationHandler.END

    step = context.user_data.get("admin_change_step", "old")

    admin_info = get_admin_by_telegram_id(update.effective_user.id)
    if not admin_info:
        await update.message.reply_text(_("admin.login_admin_inactive", lang=lang))
        return ConversationHandler.END

    if step == "old":
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT password_hash FROM admins WHERE id = ?", (admin_info["id"],)
        )
        row = cur.fetchone()
        conn.close()

        if not row or not verify_password(text, row[0]):
            await update.message.reply_text(
                _("admin.change_password_old_incorrect", lang=lang)
            )
            return ADMIN_CHANGE_PASSWORD

        context.user_data["admin_change_step"] = "new"
        await update.message.reply_text(
            _("admin.change_password_new", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.admin_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return ADMIN_CHANGE_PASSWORD

    if step == "new":
        if not validate_admin_password(text):
            await update.message.reply_text(
                _("admin.change_password_new_too_short", lang=lang)
            )
            return ADMIN_CHANGE_PASSWORD

        context.user_data["admin_change_new"] = text
        context.user_data["admin_change_step"] = "confirm"

        await update.message.reply_text(
            _("admin.change_password_confirm", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.admin_back_to_menu", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return ADMIN_CHANGE_PASSWORD

    if step == "confirm":
        new_password = context.user_data.get("admin_change_new")
        if text != new_password:
            await update.message.reply_text(
                _("admin.change_password_not_match", lang=lang)
            )
            context.user_data["admin_change_step"] = "new"
            await update.message.reply_text(
                _("admin.change_password_new", lang=lang),
                reply_markup=ReplyKeyboardMarkup(
                    [[_("buttons.admin_back_to_menu", lang=lang)]],
                    resize_keyboard=True,
                ),
            )
            return ADMIN_CHANGE_PASSWORD

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "UPDATE admins SET password_hash = ? WHERE id = ?",
            (hash_password(new_password), admin_info["id"]),
        )
        conn.commit()
        conn.close()

        context.user_data.pop("admin_change_step", None)
        context.user_data.pop("admin_change_new", None)

        await update.message.reply_text(_("admin.change_password_success", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    await admin_menu(update, context)
    return ConversationHandler.END


@require_admin
async def admin_change_username_start(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.change_username_new", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.admin_back_to_menu", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    context.user_data["admin_change_username_step"] = "new"
    return ADMIN_CHANGE_USERNAME


@require_admin
async def admin_change_username_process(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    if text == _("buttons.admin_back_to_menu", lang=lang):
        await admin_menu(update, context)
        return ConversationHandler.END

    admin_info = get_admin_by_telegram_id(update.effective_user.id)
    if not admin_info or not admin_info.get("is_active", True):
        await update.message.reply_text(_("admin.login_admin_inactive", lang=lang))
        return ConversationHandler.END

    new_username = text
    if len(new_username) < 4 or not all(
        c.isalnum() or c in "._-" for c in new_username
    ):
        await update.message.reply_text(_("admin.change_username_invalid", lang=lang))
        return ADMIN_CHANGE_USERNAME

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM admins WHERE username = ?",
        (new_username,),
    )
    row = cur.fetchone()

    if row and row[0] != admin_info["id"]:
        conn.close()
        await update.message.reply_text(_("admin.change_username_exists", lang=lang))
        return ADMIN_CHANGE_USERNAME

    cur.execute(
        "UPDATE admins SET username = ? WHERE id = ?",
        (new_username, admin_info["id"]),
    )
    conn.commit()
    conn.close()

    context.user_data.pop("admin_change_username_step", None)

    await update.message.reply_text(_("admin.change_username_success", lang=lang))
    await admin_menu(update, context)
    return ConversationHandler.END


@require_admin
async def admin_security_menu(update, context):
    lang = get_lang(context)
    keyboard = [
        [
            _("buttons.admin_menu_change_password", lang=lang),
            _("buttons.admin_menu_change_username", lang=lang),
        ],
        [
            _("buttons.admin_menu_reset_agent_password", lang=lang),
            _("buttons.admin_menu_list_admins", lang=lang),
        ],
        [
            _("buttons.admin_back_to_menu", lang=lang),
        ],
    ]
    await update.message.reply_text(
        _("admin.security_menu_title", lang=lang),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_admin
async def admin_list_admins(update, context):
    lang = get_lang(context)
    admins = get_all_admins()

    if not admins:
        await update.message.reply_text(_("admin.list_admins_empty", lang=lang))
        return

    lines = [_("admin.list_admins_title", lang=lang)]
    for row in admins:
        status = "فعال" if row["is_active"] else "غیرفعال"
        if lang == "pa":
            status = "فعال" if row["is_active"] else "غیرفعال"

        telegram = str(row["telegram_id"]) if row["telegram_id"] else "-"
        failed_attempts = row["failed_attempts"] if row["failed_attempts"] is not None else 0

        line = _(
            "admin.list_admins_row",
            lang=lang,
            id=row["id"],
            username=row["username"],
            status=status,
            telegram=telegram,
            failed_attempts=failed_attempts,
        )
        lines.append(line)

    text = "\n\n".join(lines)
    await update.message.reply_text(text)


@require_admin
async def admin_backup_db(update, context):
    lang = get_lang(context)
    db_path = "hawala.db"

    if not os.path.exists(db_path):
        await update.message.reply_text(_("admin.backup_not_found", lang=lang))
        return

    await update.message.reply_text(_("admin.backup_creating", lang=lang))

    try:
        with open(db_path, "rb") as f:
            filename = f"hawala_backup_{dt.now().strftime('%Y%m%d_%H%M%S')}.db"
            caption = _(
                "admin.backup_success_caption",
                lang=lang,
                date=dt.now().strftime("%Y-%m-%d %H:%M"),
            )
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption=caption,
            )
    except Exception as e:
        await update.message.reply_text(
            _("admin.backup_failed", lang=lang, error=str(e))
        )


@require_admin
async def admin_logout(update, context):
    user_id = update.effective_user.id

    from bot.services.database import unbind_admin_telegram_id

    unbind_admin_telegram_id(user_id)

    lang = get_lang(context)
    context.user_data.clear()

    await update.message.reply_text(
        _("admin.logout_success", lang=lang),
        reply_markup=ReplyKeyboardRemove(),
    )

    # برگشت به منوی اصلی
    from bot.handlers.start import start

    await start(update, context)


# =======================
# ➕ پیل جوړول عامل
# =======================


# تابع ایجاد عامل ساده و کارآمد
@require_admin
async def create_agent_start(update, context):
    try:
        lang = get_lang(context)
        context.user_data.clear()
        context.user_data["lang"] = lang

        await update.message.reply_text(
            _("admin.create_agent_enter_name", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.cancel", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        
        return NAME
        
    except Exception as e:
        logger.exception("Error in create_agent_start")
        await update.message.reply_text(f"❌ تېرونه په پیل کولو کې عامل: {str(e)}")
        return ConversationHandler.END





@require_admin
async def get_name(update, context):
    try:
        lang = get_lang(context)
        text = update.message.text.strip()
        if text == _("buttons.cancel", lang=lang):
            context.user_data.clear()
            context.user_data["lang"] = lang
            await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
            await admin_menu(update, context)
            return ConversationHandler.END

        context.user_data["name"] = text
        await update.message.reply_text(
            _("admin.create_agent_enter_password", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.cancel", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return PASSWORD
    except Exception as e:
        logger.exception("Error in get_name")
        await update.message.reply_text(f"❌ تېرونه په نوم لګولو کې: {str(e)}")
        return ConversationHandler.END


@require_admin
async def get_password(update, context):
    password_text = update.message.text.strip()
    lang = get_lang(context)
    
    if password_text == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    if not validate_agent_password(password_text):
        await update.message.reply_text(
            _("admin.create_agent_password_too_short", lang=lang)
        )
        return PASSWORD

    context.user_data["temp_password"] = password_text

    await update.message.reply_text(
        _("admin.create_agent_confirm_password", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return CONFIRM_PASSWORD


@require_admin
async def confirm_password(update, context):
    confirm = update.message.text.strip()
    lang = get_lang(context)
    
    if confirm == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    if confirm != context.user_data["temp_password"]:
        await update.message.reply_text(
            _("admin.create_agent_passwords_not_match", lang=lang)
        )
        return PASSWORD

    context.user_data["password"] = hash_password(confirm)
    context.user_data.pop("temp_password", None)

    await update.message.reply_text(
        _("admin.create_agent_enter_province", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return PROVINCE


@require_admin
async def get_province(update, context):
    text = update.message.text.strip()
    lang = get_lang(context)
    if text == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    context.user_data["province"] = text
    await update.message.reply_text(
        _("admin.create_agent_enter_phone", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return PHONE


@require_admin
async def get_phone(update, context):
    phone = update.message.text.strip()
    lang = get_lang(context)
    
    if phone == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM agents WHERE phone = ?", (phone,))
    exists = cur.fetchone()
    conn.close()

    if exists:
        await update.message.reply_text(
            _("admin.create_agent_phone_exists", lang=lang)
        )
        return PHONE

    context.user_data["phone"] = phone
    await update.message.reply_text(
        _("admin.create_agent_enter_tazkira", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return TAZKIRA


@require_admin
async def get_tazkira(update, context):
    tazkira = update.message.text.strip()
    lang = get_lang(context)
    
    if tazkira == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    if not tazkira.isdigit():
        await update.message.reply_text(
            _("admin.create_agent_tazkira_must_be_number", lang=lang)
        )
        return TAZKIRA

    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT id FROM agents WHERE tazkira = ?", (tazkira,))
        exists = cur.fetchone()

        conn.close()

        if exists:
            await update.message.reply_text(
                _("admin.create_agent_tazkira_exists", lang=lang)
            )
            return TAZKIRA

    except Exception:
        await update.message.reply_text(
            _("admin.create_agent_tazkira_db_error", lang=lang)
        )
        return TAZKIRA

    context.user_data["tazkira"] = tazkira

    await update.message.reply_text(
        _("admin.create_agent_enter_balance", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            [[_("buttons.cancel", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    return BALANCE


@require_admin
async def get_balance(update, context):
    try:
        balance_text = update.message.text.strip()
        lang = get_lang(context)
        
        if balance_text == _("buttons.cancel", lang=lang):
            context.user_data.clear()
            context.user_data["lang"] = lang
            await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
            await admin_menu(update, context)
            return ConversationHandler.END

        if balance_text == "0":
            balance = 0.0
        else:
            balance = float(balance_text)

        if balance < 0:
            await update.message.reply_text(
                _("admin.create_agent_balance_negative", lang=lang)
            )
            return BALANCE

    except ValueError:
        await update.message.reply_text(
            _("admin.create_agent_balance_invalid", lang=lang)
        )
        return BALANCE

    context.user_data["balance"] = balance

    keyboard = [
        ["🇦🇫 AFN", "🇺🇸 USD"],
        [_("buttons.cancel", lang=lang)],
    ]

    await update.message.reply_text(
        _("agent.choose_currency", lang=lang),
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True,
            one_time_keyboard=True,
        ),
    )
    return CURRENCY


@require_admin
async def get_currency(update, context):
    message_text = update.message.text.strip().upper()
    lang = get_lang(context)
    
    if message_text == _("buttons.cancel", lang=lang).upper():
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
        await admin_menu(update, context)
        return ConversationHandler.END

    if "AFN" in message_text:
        currency = "AFN"
    elif "USD" in message_text:
        currency = "USD"
    else:
        await update.message.reply_text("❌ فقط از دکمه‌ها استفاده کنید")
        return CURRENCY

    context.user_data["currency"] = currency

    summary = _(
        "admin.create_agent_summary",
        lang=lang,
        name=context.user_data["name"],
        province=context.user_data["province"],
        phone=context.user_data["phone"],
        tazkira=context.user_data["tazkira"],
        balance=context.user_data["balance"],
        currency=currency,
    )

    keyboard = [
        [_("buttons.confirm", lang=lang), _("buttons.cancel", lang=lang)]
    ]

    await update.message.reply_text(
        summary,
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True,
            one_time_keyboard=True,
        ),
    )

    return CONFIRM_AGENT


@require_admin
async def confirm_agent(update, context):
    text = update.message.text.strip()
    lang = get_lang(context)


    if text == _("buttons.cancel", lang=lang):
        context.user_data.clear()
        context.user_data["lang"] = lang
        await update.message.reply_text(
            _("agent.operation_cancelled", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        await admin_menu(update, context)
        return ConversationHandler.END

    if text != _("buttons.confirm", lang=lang):
        await update.message.reply_text(
            _("admin.create_agent_not_confirmed", lang=lang),
            reply_markup=ReplyKeyboardRemove(),
        )
        await admin_menu(update, context)
        return ConversationHandler.END

    try:
        required_fields = ["name", "province", "phone", "tazkira", "password", "currency", "balance"]
        missing_fields = []
        
        for field in required_fields:
            if field not in context.user_data:
                missing_fields.append(field)
        
        if missing_fields:
            await update.message.reply_text(
                _("admin.create_agent_missing_fields", lang=lang)
            )
            return ConversationHandler.END
        
        conn = get_db()
        cur = conn.cursor()

        # ثبت عامل در جدول agents
        cur.execute(
            """
            INSERT INTO agents (name, province, phone, tazkira, password_hash)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                context.user_data["name"],
                context.user_data["province"],
                context.user_data["phone"],
                context.user_data["tazkira"],
                context.user_data["password"],
            ),
        )

        agent_id = cur.lastrowid

        # ثبت موجودی در جدول balances
        cur.execute(
            """
            INSERT INTO balances (agent_id, currency, balance)
            VALUES (?, ?, ?)
            """,
            (
                agent_id,
                context.user_data["currency"],
                context.user_data["balance"],
            ),
        )

        conn.commit()
        conn.close()

        context.user_data.clear()
        context.user_data["lang"] = lang

        await update.message.reply_text(
            _(
                "admin.create_agent_success",
                lang=lang,
                agent_id=agent_id,
            ),
            reply_markup=ReplyKeyboardRemove(),
        )
        await admin_menu(update, context)
        return ConversationHandler.END

    except Exception as e:
        logger.exception("Error in confirm_agent")
        await update.message.reply_text(
            _(
                "admin.create_agent_error",
                lang=lang,
                error=str(e),
            ),
            reply_markup=ReplyKeyboardRemove(),
        )
        
        context.user_data.clear()
        context.user_data["lang"] = lang
        
        await admin_menu(update, context)
        return ConversationHandler.END


@require_admin
async def financial_report(update, context):
    lang = get_lang(context)
    await update.message.reply_text(_("admin.financial_preparing", lang=lang))

    conn = get_db()
    cur = conn.cursor()

    # آمار کل سیستم
    cur.execute("SELECT COUNT(*) FROM agents WHERE is_active = 1")
    active_agents = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM agents")
    total_agents = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM transactions")
    total_transactions = cur.fetchone()[0]

    cur.execute(
        """
        SELECT COUNT(*) FROM transactions 
        WHERE status = 'pending'
    """
    )
    pending_transactions = cur.fetchone()[0]

    cur.execute(
        """
        SELECT COUNT(*) FROM transactions 
        WHERE status = 'completed'
    """
    )
    completed_transactions = cur.fetchone()[0]

    cur.execute(
        """
        SELECT SUM(amount) FROM transactions 
        WHERE status != 'cancelled'
    """
    )
    total_amount = cur.fetchone()[0] or 0

    cur.execute(
        """
        SELECT SUM(commission) FROM transactions 
        WHERE status != 'cancelled'
    """
    )
    total_commission = cur.fetchone()[0] or 0

    # آمار بر اساس ولایت
    cur.execute(
        """
        SELECT a.province, COUNT(*) as agent_count, 
               SUM(b.balance) as total_balance
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        WHERE a.is_active = 1
        GROUP BY a.province
        ORDER BY agent_count DESC
    """
    )
    province_stats = cur.fetchall()

    # برترین عامل‌ها
    cur.execute(
        """
        SELECT a.name, a.province, COUNT(t.id) as transaction_count,
               SUM(t.commission) as total_commission
        FROM agents a
        LEFT JOIN transactions t ON a.id = t.agent_id AND t.status != 'cancelled'
        WHERE a.is_active = 1
        GROUP BY a.id, a.name, a.province
        ORDER BY transaction_count DESC
        LIMIT 5
    """
    )
    top_agents = cur.fetchall()

    conn.close()

    report = _(
        "admin.financial_report",
        lang=lang,
        active_agents=active_agents,
        total_agents=total_agents,
        total_transactions=total_transactions,
        pending_transactions=pending_transactions,
        completed_transactions=completed_transactions,
        total_amount=f"{total_amount:,.0f}",
        total_commission=f"{total_commission:,.0f}",
    )

    if province_stats:
        province_lines = []
        for province, count, balance in province_stats[:5]:
            balance_text = f"{balance:,.0f}" if balance else _(
                "common.zero", lang=lang
            )
            province_lines.append(
                _(
                    "admin.financial_province_line",
                    lang=lang,
                    province=province,
                    count=count,
                    balance=balance_text,
                )
            )
        report += "\n" + _(
            "admin.financial_province_header", lang=lang
        ) + "\n" + "\n".join(province_lines)

    if top_agents:
        top_lines = []
        for i, (name, province, count, commission) in enumerate(top_agents, 1):
            commission_text = f"{commission:,.0f}" if commission else _(
                "common.zero", lang=lang
            )
            top_lines.append(
                _(
                    "admin.financial_top_agent_line",
                    lang=lang,
                    index=i,
                    name=name,
                    province=province,
                    count=count,
                    commission=commission_text,
                )
            )
        report += "\n\n" + _(
            "admin.financial_top_agents_header", lang=lang
        ) + "\n" + "\n".join(top_lines)

    report += "\n" + _(
        "admin.financial_footer_date",
        lang=lang,
        date=dt.now().strftime("%Y/%m/%d %H:%M"),
    )

    keyboard = [
        [
            _("buttons.admin_menu_download_excel", lang=lang),
            _("admin.search_agents_back_to_menu_button", lang=lang),
        ],
    ]

    await update.message.reply_text(
        report,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


@require_admin
async def list_agents(update, context):
    lang = get_lang(context)
    if update.callback_query:
        message = update.callback_query.message
        is_callback = True
    else:
        message = update.message
        is_callback = False

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT id, name, province, phone, is_active
        FROM agents
        ORDER BY id
        """
    )
    agents_rows = cur.fetchall()
    
    cur.execute(
        """
        SELECT agent_id, balance, currency
        FROM balances
        """
    )
    balances_rows = cur.fetchall()
    conn.close()

    if not agents_rows:
        text = _("admin.list_agents_empty", lang=lang)
        if is_callback:
            await message.edit_text(text)
        else:
            await message.reply_text(text)
        return

    agent_balances = {}
    for b_agent_id, balance, currency in balances_rows:
        if b_agent_id not in agent_balances:
            agent_balances[b_agent_id] = []
        agent_balances[b_agent_id].append(f"{balance:,.0f} {currency}")

    lines = []
    active_count = 0

    for agent in agents_rows:
        agent_id, name, province, phone, is_active = agent

        if is_active:
            active_count += 1

        status = "🟢" if is_active else "🔴"

        balances_list = agent_balances.get(
            agent_id,
            [_( "admin.list_agents_zero_balance_default", lang=lang)],
        )
        balances_display = " | ".join(balances_list)

        line = _(
            "admin.list_agents_line",
            lang=lang,
            status=status,
            agent_id=f"{agent_id:03d}",
            name=name,
            province=province,
            phone=phone,
            balances=balances_display,
        )
        lines.append(line)

    current_time = dt.now().strftime("%H:%M:%S")
    inactive_count = len(agents_rows) - active_count

    header = _("admin.list_agents_header", lang=lang)
    footer = _(
        "admin.list_agents_footer",
        lang=lang,
        total=len(agents_rows),
        active=active_count,
        inactive=inactive_count,
        time=current_time,
    )

    full_text = header + "\n\n".join(lines) + footer

    keyboard = [
        [
            InlineKeyboardButton(
                _("admin.list_agents_refresh_button", lang=lang),
                callback_data="refresh_agents",
            ),
            InlineKeyboardButton(
                _("admin.list_agents_back_button", lang=lang),
                callback_data="back_to_menu",
            ),
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # ارسال/ویرایش
    if is_callback:
        await message.edit_text(
            full_text, parse_mode="Markdown", reply_markup=reply_markup
        )
    else:
        await message.reply_text(
            full_text, parse_mode="Markdown", reply_markup=reply_markup
        )


@require_admin
async def toggle_agent_start(update, context):
    try:
        lang = get_lang(context)
        context.user_data.clear()
        context.user_data["lang"] = lang
        
        await update.message.reply_text(
            _("admin.toggle_agent_enter_id", lang=lang),
            reply_markup=ReplyKeyboardMarkup(
                [[_("buttons.cancel", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        
        return TOGGLE_AGENT
        
    except Exception as e:
        logger.exception("Error in toggle_agent_start")
        await update.message.reply_text(f"❌ تېرونه په فعال/غیرفعال کولو کې: {str(e)}")
        return ConversationHandler.END


@require_admin
async def toggle_agent_by_id(update, context):
    try:
        text = update.message.text.strip()
        lang = get_lang(context)
        
        if text == _("buttons.cancel", lang=lang):
            context.user_data.clear()
            context.user_data["lang"] = lang
            await update.message.reply_text(_("agent.operation_cancelled", lang=lang))
            await admin_menu(update, context)
            return ConversationHandler.END
            
        agent_id = int(text)

        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT is_active FROM agents WHERE id = ?", (agent_id,))
        row = cur.fetchone()

        if not row:
            await update.message.reply_text(
                _("admin.toggle_agent_not_found", lang=lang)
            )
            conn.close()
            return ConversationHandler.END

        new_status = 0 if row[0] == 1 else 1

        cur.execute(
            "UPDATE agents SET is_active = ? WHERE id = ?",
            (new_status, agent_id),
        )

        conn.commit()
        conn.close()

        if new_status == 1:
            reset_failed_attempts(agent_id)

        status_text = (
            _("admin.toggle_agent_active", lang=lang)
            if new_status
            else _("admin.toggle_agent_inactive", lang=lang)
        )
        await update.message.reply_text(
            _(
                "admin.toggle_agent_status_changed",
                lang=lang,
                agent_id=agent_id,
                status=status_text,
            )
        )
        await admin_menu(update, context)
        return ConversationHandler.END

    except ValueError:
        await update.message.reply_text(
            _("admin.toggle_agent_id_must_be_number", lang=lang)
        )
        await admin_menu(update, context)
        return ConversationHandler.END
    except Exception as e:
        logger.exception("Error in toggle_agent_by_id")
        await global_error_handler(
            update,
            context,
            _("admin.toggle_agent_error", lang=lang),
        )
        await admin_menu(update, context)
        return ConversationHandler.END


@require_admin
async def list_balance_requests(update, context):
    lang = get_lang(context)
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute(
        """
        SELECT 
            br.id, br.agent_id, br.amount, br.currency, br.receipt_photo_id, br.created_at,
            a.name as agent_name
        FROM balance_requests br
        JOIN agents a ON br.agent_id = a.id
        WHERE br.status = 'pending'
        ORDER BY br.created_at ASC
        """
    )
    requests = cur.fetchall()
    conn.close()
    
    if not requests:
        await update.message.reply_text(
            _("admin.balance_requests_empty", lang=lang)
        )
        return
    
    await update.message.reply_text(
        _(
            "admin.balance_requests_header",
            lang=lang,
            count=len(requests),
        )
    )

    for req in requests:
        req_id, agent_id, amount, currency, photo_id, created_at, agent_name = req
        
        caption = _(
            "admin.balance_request_caption",
            lang=lang,
            agent_name=agent_name,
            agent_id=agent_id,
            amount=f"{amount:,.0f}",
            currency=currency,
            created_at=created_at[:16],
            request_id=req_id,
        )
        
        keyboard = [
            [
                InlineKeyboardButton(
                    _("admin.balance_request_approve_button", lang=lang),
                    callback_data=f"approve_br_{req_id}",
                ),
                InlineKeyboardButton(
                    _("admin.balance_request_reject_button", lang=lang),
                    callback_data=f"reject_br_{req_id}",
                ),
            ]
        ]
        
        try:
            await update.message.reply_photo(
                photo=photo_id,
                caption=caption,
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        except Exception as e:
            logger.error(f"Error sending request {req_id}: {e}")
            await update.message.reply_text(
                _(
                    "admin.balance_request_photo_error",
                    lang=lang,
                    request_id=req_id,
                    caption=caption,
                ),
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )


@require_admin
async def process_balance_request_callback(update, context):
    query = update.callback_query
    await query.answer()
    
    data = query.data
    if not (data.startswith("approve_br_") or data.startswith("reject_br_")):
        return

    action = "approved" if data.startswith("approve_br_") else "rejected"
    req_id = int(data.split("_")[-1])
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        cur.execute(
            "SELECT agent_id, amount, currency, status FROM balance_requests WHERE id = ?",
            (req_id,)
        )
        request = cur.fetchone()
        
        if not request:
            lang = get_lang(context)
            await query.edit_message_caption(
                _("admin.balance_request_not_found", lang=lang)
            )
            conn.close()
            return
            
        agent_id, amount, currency, status = request
        
        if status != "pending":
            lang = get_lang(context)
            await query.edit_message_caption(
                _(
                    "admin.balance_request_already_processed",
                    lang=lang,
                    status=status,
                )
            )
            conn.close()
            return

        if action == "approved":
            cur.execute(
                "SELECT balance FROM balances WHERE agent_id = ? AND currency = ?",
                (agent_id, currency)
            )
            row = cur.fetchone()
            if row:
                cur.execute(
                    "UPDATE balances SET balance = balance + ? WHERE agent_id = ? AND currency = ?",
                    (amount, agent_id, currency)
                )
            else:
                cur.execute(
                    "INSERT INTO balances (agent_id, currency, balance) VALUES (?, ?, ?)",
                    (agent_id, currency, amount)
                )
            
            cur.execute(
                "UPDATE balance_requests SET status = 'approved', processed_at = CURRENT_TIMESTAMP WHERE id = ?",
                (req_id,)
            )
            
            lang = get_lang(context)
            status_msg = _("admin.balance_request_status_approved", lang=lang)
            notif_to_agent = _(
                "admin.balance_request_notify_approved",
                lang=lang,
                amount=f"{amount:,.0f}",
                currency=currency,
            )
        else:
            cur.execute(
                "UPDATE balance_requests SET status = 'rejected', processed_at = CURRENT_TIMESTAMP WHERE id = ?",
                (req_id,)
            )
            lang = get_lang(context)
            status_msg = _("admin.balance_request_status_rejected", lang=lang)
            notif_to_agent = _(
                "admin.balance_request_notify_rejected",
                lang=lang,
                amount=f"{amount:,.0f}",
                currency=currency,
            )

        conn.commit()
        
        # اطلاع‌رسانی به عامل
        cur.execute("SELECT telegram_id FROM agents WHERE id = ?", (agent_id,))
        agent_tg = cur.fetchone()[0]
        if agent_tg:
            try:
                await context.bot.send_message(
                    chat_id=agent_tg,
                    text=notif_to_agent,
                    parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"Failed to notify agent {agent_id}: {e}")

        lang = get_lang(context)
        await query.edit_message_caption(
            _(
                "admin.balance_request_result_caption",
                lang=lang,
                caption=query.message.caption,
                status=status_msg,
            ),
            parse_mode="Markdown",
        )

    except Exception as e:
        logger.exception("Error processing balance request")
        lang = get_lang(context)
        await query.edit_message_caption(
            _(
                "admin.balance_request_result_error",
                lang=lang,
                caption=query.message.caption,
                error=str(e),
            )
        )
    finally:
        conn.close()

@require_admin
async def handle_agents_callback(update, context):
    query = update.callback_query
    await query.answer()

    data = query.data

    lang = get_lang(context)

    if data == "refresh_agents":
        await query.edit_message_text(
            _("admin.list_agents_refreshing", lang=lang),
            reply_markup=None,
        )
        await list_agents(update, context)

    elif data == "back_to_menu":
        await query.edit_message_text(
            _("admin.list_agents_backing_to_menu", lang=lang),
            reply_markup=None,
        )
        fake_update = Update(update_id=update.update_id, message=query.message)

        await admin_menu(fake_update, context)  # ✅ به منوی اصلی


# =======================
# 🔍 پرمختللې لټون عاملانو
# =======================


@require_admin
async def search_agents(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.search_agents_menu", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [
                [
                    _("admin.search_agents_by_name_button", lang=lang),
                ],
                [
                    _("admin.search_agents_by_province_button", lang=lang),
                ],
                [
                    _("admin.search_agents_by_phone_button", lang=lang),
                ],
                [
                    _("admin.search_agents_only_active_button", lang=lang),
                ],
                [
                    _("admin.search_agents_only_inactive_button", lang=lang),
                ],
                [
                    _("admin.search_agents_back_to_menu_button", lang=lang),
                ],
            ],
            resize_keyboard=True,
        )
    )


@require_admin
async def admin_search_tx_start(update, context):
    lang = get_lang(context)
    keyboard = [
        [
            _("admin.tx_report_range_today", lang=lang),
            _("admin.tx_report_range_7d", lang=lang),
        ],
        [
            _("admin.tx_report_range_30d", lang=lang),
            _("admin.tx_report_range_90d", lang=lang),
        ],
        [
            _("admin.tx_report_range_custom", lang=lang),
        ],
        [
            _("admin.search_agents_back_to_menu_button", lang=lang),
        ],
    ]

    await update.message.reply_text(
        _("admin.tx_report_menu_title", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )
    context.user_data["tx_report_state"] = "select_range"
    return ADMIN_SEARCH_TX


@require_admin
async def admin_search_tx_process(update, context):
    lang = get_lang(context)
    text = update.message.text.strip()

    back_label = _("admin.search_agents_back_to_menu_button", lang=lang)
    custom_label = _("admin.tx_report_range_custom", lang=lang)
    today_label = _("admin.tx_report_range_today", lang=lang)
    week_label = _("admin.tx_report_range_7d", lang=lang)
    month_label = _("admin.tx_report_range_30d", lang=lang)
    quarter_label = _("admin.tx_report_range_90d", lang=lang)

    state = context.user_data.get("tx_report_state", "select_range")

    if text == back_label:
        context.user_data.pop("tx_report_state", None)
        context.user_data.pop("tx_report_start", None)
        context.user_data.pop("tx_report_end", None)
        await admin_menu(update, context)
        return ConversationHandler.END

    if state == "select_range":
        if text == today_label:
            end_date = dt.now().date()
            start_date = end_date
        elif text == week_label:
            end_date = dt.now().date()
            start_date = end_date - datetime.timedelta(days=7)
        elif text == month_label:
            end_date = dt.now().date()
            start_date = end_date - datetime.timedelta(days=30)
        elif text == quarter_label:
            end_date = dt.now().date()
            start_date = end_date - datetime.timedelta(days=90)
        elif text == custom_label:
            context.user_data["tx_report_state"] = "custom_start"
            await update.message.reply_text(
                _("admin.tx_report_enter_start", lang=lang),
                reply_markup=ReplyKeyboardMarkup(
                    [[back_label]],
                    resize_keyboard=True,
                ),
            )
            return ADMIN_SEARCH_TX
        else:
            await update.message.reply_text(
                _("admin.tx_report_menu_title", lang=lang),
                parse_mode="Markdown",
            )
            return ADMIN_SEARCH_TX

        context.user_data["tx_report_start"] = start_date
        context.user_data["tx_report_end"] = end_date
        context.user_data["tx_report_state"] = "select_status"

        status_keyboard = [
            [
                _("admin.tx_report_status_all", lang=lang),
            ],
            [
                _("admin.tx_report_status_completed", lang=lang),
                _("admin.tx_report_status_pending", lang=lang),
            ],
            [
                _("admin.tx_report_status_cancelled", lang=lang),
            ],
            [
                back_label,
            ],
        ]

        await update.message.reply_text(
            _("admin.tx_report_status_menu_title", lang=lang),
            reply_markup=ReplyKeyboardMarkup(status_keyboard, resize_keyboard=True),
        )
        return ADMIN_SEARCH_TX

    if state == "custom_start":
        try:
            start_date = dt.strptime(text, "%Y-%m-%d").date()
        except Exception:
            await update.message.reply_text(
                _("admin.tx_report_invalid_date", lang=lang)
            )
            return ADMIN_SEARCH_TX
        context.user_data["tx_report_start"] = start_date
        context.user_data["tx_report_state"] = "custom_end"
        await update.message.reply_text(_("admin.tx_report_enter_end", lang=lang))
        return ADMIN_SEARCH_TX

    if state == "custom_end":
        try:
            end_date = dt.strptime(text, "%Y-%m-%d").date()
        except Exception:
            await update.message.reply_text(
                _("admin.tx_report_invalid_date", lang=lang)
            )
            return ADMIN_SEARCH_TX

        start_date = context.user_data.get("tx_report_start")
        if not start_date or end_date < start_date:
            await update.message.reply_text(
                _("admin.tx_report_invalid_range", lang=lang)
            )
            return ADMIN_SEARCH_TX

        context.user_data["tx_report_start"] = start_date
        context.user_data["tx_report_end"] = end_date
        context.user_data["tx_report_state"] = "select_status"

        status_keyboard = [
            [
                _("admin.tx_report_status_all", lang=lang),
            ],
            [
                _("admin.tx_report_status_completed", lang=lang),
                _("admin.tx_report_status_pending", lang=lang),
            ],
            [
                _("admin.tx_report_status_cancelled", lang=lang),
            ],
            [
                back_label,
            ],
        ]

        await update.message.reply_text(
            _("admin.tx_report_status_menu_title", lang=lang),
            reply_markup=ReplyKeyboardMarkup(status_keyboard, resize_keyboard=True),
        )
        return ADMIN_SEARCH_TX

    if state == "select_status":
        all_label = _("admin.tx_report_status_all", lang=lang)
        completed_label = _("admin.tx_report_status_completed", lang=lang)
        pending_label = _("admin.tx_report_status_pending", lang=lang)
        cancelled_label = _("admin.tx_report_status_cancelled", lang=lang)

        valid_labels = {all_label, completed_label, pending_label, cancelled_label}

        if text not in valid_labels:
            await update.message.reply_text(
                _("admin.tx_report_status_menu_title", lang=lang)
            )
            return ADMIN_SEARCH_TX

        status_filter = None
        status_label = all_label

        if text == completed_label:
            status_filter = "completed"
            status_label = completed_label
        elif text == pending_label:
            status_filter = "pending"
            status_label = pending_label
        elif text == cancelled_label:
            status_filter = "cancelled"
            status_label = cancelled_label

        start_date = context.user_data.get("tx_report_start")
        end_date = context.user_data.get("tx_report_end")

        if not start_date or not end_date:
            context.user_data["tx_report_state"] = "select_range"
            return await admin_search_tx_start(update, context)

        await _send_admin_tx_excel(
            update, context, lang, start_date, end_date, status_filter, status_label
        )

        status_keyboard = [
            [
                all_label,
            ],
            [
                completed_label,
                pending_label,
            ],
            [
                cancelled_label,
            ],
            [
                back_label,
            ],
        ]

        await update.message.reply_text(
            _("admin.tx_report_status_menu_title", lang=lang),
            reply_markup=ReplyKeyboardMarkup(status_keyboard, resize_keyboard=True),
        )
        return ADMIN_SEARCH_TX

    context.user_data["tx_report_state"] = "select_range"
    return await admin_search_tx_start(update, context)


async def _send_admin_tx_excel(
    update, context, lang, start_date, end_date, status_filter, status_label
):
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    conn = get_db()
    cur = conn.cursor()
    base_query = """
        SELECT 
            t.transaction_code,
            t.sender_name,
            t.receiver_name,
            t.receiver_tazkira,
            t.amount,
            t.currency,
            t.commission,
            t.status,
            t.created_at,
            t.completed_at,
            a_sender.name as sender_agent,
            a_receiver.name as receiver_agent,
            a_receiver.province as receiver_province
        FROM transactions t
        JOIN agents a_sender ON t.agent_id = a_sender.id
        JOIN agents a_receiver ON t.receiver_agent_id = a_receiver.id
        WHERE DATE(t.created_at) BETWEEN ? AND ?
    """

    params = [start_str, end_str]

    if status_filter in {"pending", "completed", "cancelled"}:
        base_query += " AND t.status = ?"
        params.append(status_filter)

    base_query += " ORDER BY t.created_at DESC"

    cur.execute(base_query, params)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text(
            _("admin.tx_report_no_transactions", lang=lang)
        )
        return

    tx_code_col = _("agent.excel_col_tx_code", lang=lang)
    sender_name_col = _("agent.excel_col_sender_name", lang=lang)
    receiver_name_col = _("agent.excel_col_receiver_name", lang=lang)
    receiver_tazkira_col = _("agent.excel_col_receiver_tazkira", lang=lang)
    amount_col = _("agent.excel_col_amount", lang=lang)
    currency_col = _("agent.excel_col_currency", lang=lang)
    commission_col = _("agent.excel_col_commission", lang=lang)
    status_col = _("agent.excel_col_status", lang=lang)
    created_at_col = _("agent.excel_col_created_at", lang=lang)
    completed_at_col = _("agent.excel_col_completed_at", lang=lang)
    receiver_agent_col = _("agent.excel_col_receiver_agent", lang=lang)
    receiver_province_col = _("agent.excel_col_receiver_province", lang=lang)

    df = pd.DataFrame(
        rows,
        columns=[
            tx_code_col,
            sender_name_col,
            receiver_name_col,
            receiver_tazkira_col,
            amount_col,
            currency_col,
            commission_col,
            status_col,
            created_at_col,
            completed_at_col,
            _("agent.excel_agent_name", lang=lang),
            receiver_agent_col,
            receiver_province_col,
        ],
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = _("agent.excel_sheet_transactions", lang=lang)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        status_col_name = _("agent.excel_col_status", lang=lang)
        amount_col_name = _("agent.excel_col_amount", lang=lang)
        commission_col_name = _("agent.excel_col_commission", lang=lang)

        summary_type_col = _("agent.excel_summary_type", lang=lang)
        summary_value_col = _("agent.excel_summary_value", lang=lang)

        summary_rows = [
            [
                _("agent.excel_summary_total_txs", lang=lang),
                len(df),
            ],
            [
                _("agent.excel_summary_pending", lang=lang),
                len(df[df[status_col_name] == "pending"]),
            ],
            [
                _("agent.excel_summary_completed", lang=lang),
                len(df[df[status_col_name] == "completed"]),
            ],
            [
                _("agent.excel_summary_cancelled", lang=lang),
                len(df[df[status_col_name] == "cancelled"]),
            ],
            [
                _("agent.excel_summary_total_amount", lang=lang),
                float(df[amount_col_name].sum() or 0),
            ],
            [
                _("agent.excel_summary_total_commission", lang=lang),
                float(df[commission_col_name].sum() or 0),
            ],
        ]

        summary_df = pd.DataFrame(
            summary_rows,
            columns=[summary_type_col, summary_value_col],
        )

        summary_sheet_name = _("agent.excel_sheet_summary", lang=lang)
        summary_df.to_excel(
            writer,
            sheet_name=summary_sheet_name,
            index=False,
        )

        workbook = writer.book

        if summary_sheet_name in workbook.sheetnames:
            sheet_summary = workbook[summary_sheet_name]
            chart = BarChart()
            chart.title = _("agent.excel_summary_chart_title", lang=lang)
            chart.y_axis.title = summary_value_col
            chart.x_axis.title = ""
            data = Reference(
                sheet_summary,
                min_col=2,
                max_col=2,
                min_row=1,
                max_row=sheet_summary.max_row,
            )
            chart.add_data(data, titles_from_data=True)
            cats = Reference(
                sheet_summary,
                min_col=1,
                min_row=2,
                max_row=sheet_summary.max_row,
            )
            chart.set_categories(cats)
            sheet_summary.add_chart(chart, "D2")

    output.seek(0)

    filename_prefix = _("agent.excel_filename_prefix", lang=lang)
    filename = f"{filename_prefix}_{start_str}_{end_str}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    caption = _(
        "admin.tx_report_excel_caption",
        lang=lang,
        start=start_str,
        end=end_str,
    )

    if status_label:
        caption += "\n" + _(
            "admin.tx_report_status_label",
            lang=lang,
            status=status_label,
        )

    await update.message.reply_document(
        document=output,
        filename=filename,
        caption=caption,
        parse_mode="Markdown",
    )


@require_admin
async def search_by_name(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.search_by_name_prompt", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [[_("admin.search_back_button", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    context.user_data["search_type"] = "name"


@require_admin
async def search_by_province(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.search_by_province_prompt", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [[_("admin.search_back_button", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    context.user_data["search_type"] = "province"


@require_admin
async def search_by_phone(update, context):
    lang = get_lang(context)
    await update.message.reply_text(
        _("admin.search_by_phone_prompt", lang=lang),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [[_("admin.search_back_button", lang=lang)]],
            resize_keyboard=True,
        ),
    )
    context.user_data["search_type"] = "phone"


@require_admin
async def execute_search(update, context):
    search_term = update.message.text.strip()
    search_type = context.user_data.get("search_type")
    lang = get_lang(context)
    
    if search_term == _("admin.search_back_button", lang=lang):
        await search_agents(update, context)
        return
    
    if not search_type:
        await search_agents(update, context)
        return
    
    conn = get_db()
    cur = conn.cursor()
    
    query = """
        SELECT a.id, a.name, a.province, a.phone, a.is_active,
               b.balance, b.currency,
               COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        LEFT JOIN transactions t ON a.id = t.agent_id
    """
    params = []
    
    if search_type == "name":
        query += " WHERE a.name LIKE ?"
        params.append(f"%{search_term}%")
    elif search_type == "province":
        query += " WHERE a.province LIKE ?"
        params.append(f"%{search_term}%")
    elif search_type == "phone":
        query += " WHERE a.phone LIKE ?"
        params.append(f"%{search_term}%")
    
    query += " GROUP BY a.id ORDER BY a.name"
    
    cur.execute(query, params)
    results = cur.fetchall()
    conn.close()
    
    if not results:
        await update.message.reply_text(
            _(
                "admin.search_agents_not_found",
                lang=lang,
                term=search_term,
            ),
            reply_markup=ReplyKeyboardMarkup(
                [[_("admin.search_back_button", lang=lang)]],
                resize_keyboard=True,
            ),
        )
        return
    
    report = _(
        "admin.search_agents_results_header",
        lang=lang,
        count=len(results),
    )
    
    for agent_id, name, province, phone, is_active, balance, currency, transaction_count in results:
        status = "🟢" if is_active else "🔴"
        balance_display = (
            f"{balance:,.0f}"
            if balance
            else _("common.zero", lang=lang)
        )
        currency_display = currency if currency else _("common.afn", lang=lang)
        
        report += _(
            "admin.search_agents_result_line",
            lang=lang,
            status=status,
            agent_id=f"{agent_id:03d}",
            name=name,
            province=province,
            phone=phone,
            balance=balance_display,
            currency=currency_display,
            tx_count=transaction_count,
        )
    
    report += _("admin.search_agents_results_footer", lang=lang)
    
    await update.message.reply_text(
        report,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("admin.search_new_button", lang=lang)],
                [_("admin.search_agents_back_to_menu_button", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )
    
    context.user_data.pop("search_type", None)


# د ادمین لپاره ساده لټون تابع
async def admin_search_handler(update, context):
    search_term = update.message.text.strip()
    search_type = context.user_data.get("search_type")
    lang = get_lang(context)
    
    # که که لټون حالت کې نه یو، بیرغرد
    if not search_type:
        return
    
    if search_term == _("admin.search_back_button", lang=lang):
        await search_agents(update, context)
        return
    
    conn = get_db()
    cur = conn.cursor()
    
    query = """
        SELECT a.id, a.name, a.province, a.phone, a.is_active,
               b.balance, b.currency,
               COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        LEFT JOIN transactions t ON a.id = t.agent_id
    """
    params = []
    
    if search_type == "name":
        query += " WHERE a.name LIKE ?"
        params.append(f"%{search_term}%")
    elif search_type == "province":
        query += " WHERE a.province LIKE ?"
        params.append(f"%{search_term}%")
    elif search_type == "phone":
        query += " WHERE a.phone LIKE ?"
        params.append(f"%{search_term}%")
    
    query += " GROUP BY a.id ORDER BY a.name"
    
    try:
        cur.execute(query, params)
        results = cur.fetchall()
        conn.close()
        
        if not results:
            await update.message.reply_text(
                _(
                    "admin.search_agents_not_found_with_type",
                    lang=lang,
                    term=search_term,
                    search_type=search_type,
                ),
                reply_markup=ReplyKeyboardMarkup(
                    [[_("admin.search_back_button", lang=lang)]],
                    resize_keyboard=True,
                ),
            )
            return
        
        report = _(
            "admin.search_agents_results_header",
            lang=lang,
            count=len(results),
        )
        
        for agent_id, name, province, phone, is_active, balance, currency, transaction_count in results:
            status = "🟢" if is_active else "🔴"
            balance_display = (
                f"{balance:,.0f}"
                if balance
                else _("common.zero", lang=lang)
            )
            currency_display = (
                currency if currency else _("common.afn", lang=lang)
            )
            
            report += _(
                "admin.search_agents_result_line",
                lang=lang,
                status=status,
                agent_id=f"{agent_id:03d}",
                name=name,
                province=province,
                phone=phone,
                balance=balance_display,
                currency=currency_display,
                tx_count=transaction_count,
            )
        
        report += _("admin.search_agents_results_footer", lang=lang)
        
        await update.message.reply_text(
            report,
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [_("admin.search_new_button", lang=lang)],
                    [_("admin.search_agents_back_to_menu_button", lang=lang)],
                ],
                resize_keyboard=True,
            ),
        )
        
        context.user_data.pop("search_type", None)
        
    except Exception as e:
        logger.exception("Error in admin search")
        await update.message.reply_text(
            _("admin.search_error", lang=lang)
        )
        context.user_data.pop("search_type", None)


@require_admin
async def filter_active_agents(update, context):
    lang = get_lang(context)
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT a.id, a.name, a.province, a.phone,
               b.balance, b.currency,
               COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        LEFT JOIN transactions t ON a.id = t.agent_id
        WHERE a.is_active = 1
        GROUP BY a.id
        ORDER BY a.name
    """)
    results = cur.fetchall()
    conn.close()
    
    if not results:
        await update.message.reply_text(
            _("admin.filter_active_empty", lang=lang)
        )
        return
    
    report = _(
        "admin.filter_active_header",
        lang=lang,
        count=len(results),
    )
    
    for agent_id, name, province, phone, balance, currency, transaction_count in results:
        balance_display = (
            f"{balance:,.0f}"
            if balance
            else _("common.zero", lang=lang)
        )
        currency_display = currency if currency else _("common.afn", lang=lang)
        
        report += _(
            "admin.filter_active_line",
            lang=lang,
            agent_id=f"{agent_id:03d}",
            name=name,
            province=province,
            phone=phone,
            balance=balance_display,
            currency=currency_display,
            tx_count=transaction_count,
        )
    
    report += _("admin.filter_active_footer", lang=lang)
    
    await update.message.reply_text(
        report,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("admin.search_new_button", lang=lang)],
                [_("admin.search_agents_back_to_menu_button", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )


@require_admin
async def filter_inactive_agents(update, context):
    lang = get_lang(context)
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT a.id, a.name, a.province, a.phone,
               b.balance, b.currency,
               COUNT(t.id) as transaction_count
        FROM agents a
        LEFT JOIN balances b ON a.id = b.agent_id
        LEFT JOIN transactions t ON a.id = t.agent_id
        WHERE a.is_active = 0
        GROUP BY a.id
        ORDER BY a.name
    """)
    results = cur.fetchall()
    conn.close()
    
    if not results:
        await update.message.reply_text(
            _("admin.filter_inactive_empty", lang=lang)
        )
        return
    
    report = _(
        "admin.filter_inactive_header",
        lang=lang,
        count=len(results),
    )
    
    for agent_id, name, province, phone, balance, currency, transaction_count in results:
        balance_display = (
            f"{balance:,.0f}"
            if balance
            else _("common.zero", lang=lang)
        )
        currency_display = currency if currency else _("common.afn", lang=lang)
        
        report += _(
            "admin.filter_inactive_line",
            lang=lang,
            agent_id=f"{agent_id:03d}",
            name=name,
            province=province,
            phone=phone,
            balance=balance_display,
            currency=currency_display,
            tx_count=transaction_count,
        )
    
    report += _("admin.filter_inactive_footer", lang=lang)
    
    await update.message.reply_text(
        report,
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(
            [
                [_("admin.search_new_button", lang=lang)],
                [_("admin.search_agents_back_to_menu_button", lang=lang)],
            ],
            resize_keyboard=True,
        ),
    )
